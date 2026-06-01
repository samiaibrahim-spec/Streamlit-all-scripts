#!/usr/bin/env python3
"""
WoW Performance Update Report Generator — Streamlit App

Converts SA360 CSV/XLSX export into formatted Excel report.

Usage: streamlit run wow_report_app.py
"""

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import os

# =============================================================================
# CONFIGURATION
# =============================================================================

METRICS = [
    'Impr.', 'Clicks', 'Cost', 'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New',
    # TODO: Address Capture is consistently ~12-200% higher than the reference report across all
    # sections. This is likely a conversion window or attribution setting difference — the manual
    # SA360 pull may be using click-through only while this export includes view-through conversions.
    # Verify the SA360 export settings match the manual pull: check Attribution model and
    # Conversion window under Campaign > Columns > Conversions before regenerating.
    'Address Capture', 'Begin Checkout',
    #'Main Sales Number', 'Contact Us Page',
    'Quality Sales Call - AN',
    'Total Conversions - VBB', 'Total Conversion Value - VBB',
    'Chat Initiation - Order Services',
]

TOTAL_ACTIONS_COMPONENTS = [
    'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New',
    'Chat Initiation - Order Services',
    'Quality Sales Call - AN',
]

# Labels used by all other specific campaign buckets — used to define
# NC Non-Testing as "all NC campaigns NOT in any of these groups"
NC_SPECIFIC_LABELS = [
    '2026 VBB Google Campaigns',
    'CBB NB Internet STD Campaigns',
    '2026 UpMarket Campaigns',
    # REMOVED: Nonbrand Consolidation 3.19.26 — reference report does not exclude these
    # campaigns from NC Non-Testing. Removing this so the script matches reference behavior.
    # Note: Nonbrand Consolidation 4.6.26 (newer label) was never in this list either.
    '2026 CBB NB Remaining Google Campaigns',
    'MSFT NB Max Clicks Campaigns',
]

# Campaign name substrings to exclude from NC Non-Testing — mirrors the SA360
# "Campaign name does not contain" filters visible in the reference report filter view.
NC_EXCLUDE_CAMPAIGN_PATTERNS = [
    'discovery',
    'master',
    'midlife',
]

# Account names to exclude from NC Non-Testing — mirrors "Account name does not contain rapidscale"
NC_EXCLUDE_ACCOUNTS = [
    'rapidscale',
]

TABLES = [
    ('All SEM', {}, 'standard'),
    ('Brand SEM', {'Brand/NB': 'Brand'}, 'standard'),
    ('Nonbrand SEM', {'Brand/NB': 'NB'}, 'standard'),
    ('NC VBB Campaigns', {'Labels on Campaign: Directly Applied': '2026 VBB Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    # TODO: verify the exact label string for CBB NB Internet Campaigns in your current SA360 export
    ('NC CBB NB Internet Campaigns', {'Labels on Campaign: Directly Applied': 'CBB NB Internet STD Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC UpMarket Campaigns', {'Labels on Campaign: Directly Applied': '2026 UpMarket Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    # REMOVED: NB Consolidated Campaigns — this section does not exist in the reference report.
    # The Nonbrand Consolidation 3.19.26 label is still excluded from NC Non-Testing via NC_SPECIFIC_LABELS.
    # FIX: was using 'Test Segment' column which is not in the SA360 export — switched to label-based filter
    # TODO: verify the label string matches what's in your current SA360 export (e.g. 'MSFT CBB NB Campaigns Feb 26')
    #('MSFT CBB NB Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT CBB NB Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Google Campaigns', {'Labels on Campaign: Directly Applied': '2026 CBB NB Remaining Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC Max Clicks NB MSFT Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT NB Max Clicks Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    # NC Non-Testing: negative exclusion approach matching the SA360 filter view —
    # all NC campaigns that don't carry any specific bucket label, don't contain
    # discovery/master/midlife in the campaign name, and aren't in the rapidscale account.
    ('NC Non-Testing Campaigns', {
        'Customer Type': 'NC',
        'exclude_labels': NC_SPECIFIC_LABELS,
        'exclude_campaign_patterns': NC_EXCLUDE_CAMPAIGN_PATTERNS,
        'exclude_accounts': NC_EXCLUDE_ACCOUNTS,
    }, 'vbb'),
]

STANDARD_COLS = [
    ('Date Range', None), ('Tactic', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    #('Main Sales Number', 'Main Sales Number'), ('Contact Us Page', 'Contact Us Page'),
    # NOTE: Quality Sales Calls and Chat Initiation use their real metric keys so col_map
    # registers them correctly for the total_actions formula. They are zeroed out during
    # write via the 'standard' table type check in _write_data_row — do NOT change these
    # to suppress_zero or any non-metric key or total_actions will silently drop them.
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Conversions - VBB', 'Total Conversions - VBB'),
    ('Total Conversion Value - VBB', 'Total Conversion Value - VBB'),
    ('Total Actions', 'total_actions'),
    ('Cost per Action', 'cpactions'),
]

VBB_COLS = [
    ('Date Range', None), ('Campaign', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Conversions - VBB', 'Total Conversions - VBB'),
    ('Total Conversion Value - VBB', 'Total Conversion Value - VBB'),
    ('Total Actions', 'total_actions'),
    # FIX: was 'cost per action' (lowercase) which didn't match the 'cpactions' handler
    # in _write_data_row, causing Cost per Action to always write 0 for VBB tables
    ('Cost per Action', 'cpactions'),
]

# Column aliases — maps the standard name to known alternatives
COLUMN_ALIASES = {
    'Campaign': ['Campaign', 'Campaign Name', 'campaign'],
    'Week (Mon to Sun)': [
        'Week (Mon to Sun)', 'Week (Mon - Sun)',
        'Week (mon to sun)', 'Week (mon - sun)',
        'Week', 'week',
    ],
    'Cost': ['Cost', 'Spend', 'cost', 'spend'],
    'Clicks': ['Clicks', 'clicks'],
    'Impr.': ['Impr.', 'Impressions', 'Impr', 'impressions', 'impr.'],
    'Labels on Campaign: Directly Applied': [
        'Labels on Campaign: Directly Applied',
        'Labels on campaign: Directly applied',
        'Labels',
        'Campaign Labels',
    ],
}

# =============================================================================
# CLASSIFICATION FUNCTIONS
# =============================================================================

def classify_customer_type(campaign):
    if pd.isna(campaign):
        return 'NC'
    return 'CC' if '_CC_' in campaign else 'NC'

def classify_brand_nb(campaign):
    if pd.isna(campaign):
        return 'Brand'
    if '_Nonbr_' in campaign:
        return 'NB'
    return 'Brand'

def add_classifications(df):
    df = df.copy()
    df['Customer Type'] = df['Campaign'].apply(classify_customer_type)
    df['Brand/NB'] = df['Campaign'].apply(classify_brand_nb)
    labels_col = 'Labels on Campaign: Directly Applied'
    if labels_col not in df.columns:
        df[labels_col] = ''
    return df

# =============================================================================
# DATA LOADING AND PROCESSING
# =============================================================================

def normalize_columns(df):
    """Rename columns to standard names using COLUMN_ALIASES."""
    df.columns = df.columns.str.strip()
    rename_map = {}
    for standard_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns and alias != standard_name:
                rename_map[alias] = standard_name
                break
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

def check_required_columns(df):
    """Check if required columns are present. Returns (ok, missing_list)."""
    required = ['Campaign', 'Week (Mon to Sun)', 'Cost', 'Clicks', 'Impr.']
    missing = [c for c in required if c not in df.columns]
    return len(missing) == 0, missing

def load_file(uploaded_file):
    """Load CSV or Excel file, normalize columns, and validate."""
    name = uploaded_file.name.lower()

    # --- Excel ---
    if name.endswith(('.xlsx', '.xls')):
        for skip in [2, 0, 1, 3]:
            try:
                df = pd.read_excel(uploaded_file, skiprows=skip)
                df = normalize_columns(df)
                ok, missing = check_required_columns(df)
                if ok:
                    return df, None
            except Exception:
                continue
            finally:
                uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, skiprows=2)
        uploaded_file.seek(0)
        df = normalize_columns(df)
        _, missing = check_required_columns(df)
        return None, f"Could not find required columns in Excel file. Missing: {missing}. Found: {list(df.columns[:15])}"

    # --- CSV / TSV ---
    attempts = [
        ('utf-16', '\t', 2), ('utf-16-le', '\t', 2),
        ('utf-8', ',', 2), ('utf-8', '\t', 2), ('utf-8', ';', 2),
        ('utf-8', ',', 0), ('utf-8', '\t', 0), ('utf-8', ';', 0),
        ('latin-1', ',', 2), ('latin-1', '\t', 2), ('latin-1', ';', 2),
        ('latin-1', ',', 0), ('latin-1', '\t', 0), ('latin-1', ';', 0),
    ]
    last_columns = []
    for enc, sep, skip in attempts:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=enc, sep=sep, skiprows=skip)
            df = normalize_columns(df)
            ok, missing = check_required_columns(df)
            if ok:
                return df, None
            last_columns = list(df.columns[:15])
        except Exception:
            continue
    return None, f"Could not parse CSV. Last columns found: {last_columns}"

def clean_numerics(df):
    for col in METRICS:
        if col in df.columns:
            df[col] = df[col].replace(['--', ' --', '- ', '-'], 0)
            df[col] = df[col].replace(r'[\$,US]', '', regex=True)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df

def aggregate(df, filters):
    filt = df.copy()
    labels_col = 'Labels on Campaign: Directly Applied'

    for col, val in filters.items():
        if col == 'exclude_labels':
            # Exclude rows whose label contains any of the listed strings
            if labels_col in filt.columns:
                for label in val:
                    filt = filt[~filt[labels_col].astype(str).str.contains(label, na=False)]
        elif col == 'exclude_campaign_patterns':
            # Exclude rows where campaign name contains any of the listed substrings (case-insensitive)
            # Mirrors SA360 "Campaign name does not contain X" filters
            if 'Campaign' in filt.columns:
                for pattern in val:
                    filt = filt[~filt['Campaign'].astype(str).str.contains(pattern, case=False, na=False)]
        elif col == 'exclude_accounts':
            # Exclude rows where account name contains any of the listed substrings (case-insensitive)
            # Mirrors SA360 "Account name does not contain rapidscale" filter
            account_col = next((c for c in filt.columns if 'account' in c.lower()), None)
            if account_col:
                for acct in val:
                    filt = filt[~filt[account_col].astype(str).str.contains(acct, case=False, na=False)]
        elif col not in filt.columns:
            continue
        elif col == labels_col:
            filt = filt[filt[col].astype(str).str.contains(val, na=False)]
        else:
            filt = filt[filt[col] == val]

    weeks = sorted(filt['Week (Mon to Sun)'].dropna().unique())
    if len(weeks) < 2:
        raise ValueError(f"Need 2+ weeks, found {len(weeks)}")

    result = {}
    for label, week in [('current', weeks[-1]), ('prior', weeks[-2])]:
        wdata = filt[filt['Week (Mon to Sun)'] == week]
        agg = {c: wdata[c].sum() if c in wdata.columns else 0 for c in METRICS}
        agg['week'] = week
        result[label] = agg
    return result

# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def fmt_date(week):
    if pd.isna(week):
        return ''
    try:
        s = pd.to_datetime(week)
        e = s + pd.Timedelta(days=6)
        return f"{s.month}/{s.day}-{e.month}/{e.day}"
    except Exception:
        return str(week)

def _build_column_map(cols):
    col_map = {}
    for i, (cname, key) in enumerate(cols, start=2):
        if key is not None:
            col_map[key] = get_column_letter(i)
    return col_map


# Keys that should be zeroed out in standard (tactic-level) rows.
# These must still use their real metric keys in STANDARD_COLS so col_map registers
# them correctly for the total_actions formula — suppression happens here at write time.
STANDARD_SUPPRESS_ZERO_KEYS = {
    'Quality Sales Call - AN',
    'Chat Initiation - Order Services',
}

def _safe_num(v):
    """Return float(v) or 0 if v is missing/non-numeric."""
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0

def _write_data_row(ws, row, cols, col_map, agg_data, table_type='vbb'):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Pre-compute derived values as real numbers — avoids Excel formula
    # caching issue where openpyxl-written formulas have no cached value
    # and some Excel/Sheets versions show blank cells instead of calculating.
    clicks = _safe_num(agg_data.get('Clicks', 0))
    cost   = _safe_num(agg_data.get('Cost', 0))
    impr   = _safe_num(agg_data.get('Impr.', 0))

    # Total actions: standard tables suppress QSC and Chat so only sum eCom + Lead
    if table_type == 'standard':
        ta = (_safe_num(agg_data.get('CB eCom Order Tag - New', 0)) +
              _safe_num(agg_data.get('CB General Lead Form Submission - New', 0)))
    else:
        ta = sum(_safe_num(agg_data.get(k, 0)) for k in TOTAL_ACTIONS_COMPONENTS)

    cpc = (cost / clicks) if clicks else 0.0
    ctr = (clicks / impr) if impr   else 0.0
    cpa = (cost / ta)     if ta     else 0.0

    for i, (cname, key) in enumerate(cols[2:], start=4):
        c = ws.cell(row=row, column=i)
        c.border = border
        if key == 'cpc':
            c.value = round(cpc, 2)
            c.number_format = '$#,##0.00'
        elif key == 'ctr':
            c.value = round(ctr, 6)
            c.number_format = '0.00%'
        elif key == 'suppress_blank':
            c.value = ''
        elif key == 'total_actions':
            c.value = round(ta, 2) if ta != int(ta) else int(ta)
            c.number_format = '#,##0'
        elif key == 'cpactions':
            c.value = round(cpa, 2)
            c.number_format = '$#,##0.00'
        elif key is None:
            c.value = ''
        else:
            if table_type == 'standard' and key in STANDARD_SUPPRESS_ZERO_KEYS:
                c.value = 0
                c.number_format = '#,##0'
            else:
                v = agg_data.get(key, 0)
                c.value = int(v) if isinstance(v, float) and v == int(v) else v
                if 'Cost' in cname or 'Value' in cname:
                    c.number_format = '$#,##0.00'

def create_report(df):
    """Generate the Excel report and return it as a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WoW Performance Update"

    hfont = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    pctfill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    valign = Alignment(vertical='center')

    row = 1
    skipped = []

    for tname, filters, ctype in TABLES:
        try:
            agg = aggregate(df, filters)
        except Exception as e:
            skipped.append((tname, str(e)))
            continue

        cols = STANDARD_COLS if ctype == 'standard' else VBB_COLS
        col_map = _build_column_map(cols)

        # Header row
        for i, (cname, _) in enumerate(cols, start=2):
            c = ws.cell(row=row, column=i, value=cname)
            c.font = hfont
            c.border = border
        row += 1

        # Prior week row
        ws.cell(row=row, column=2, value=fmt_date(agg['prior']['week'])).border = border
        tactic_cell = ws.cell(row=row, column=3, value=tname)
        tactic_cell.border = border
        tactic_cell.alignment = valign
        _write_data_row(ws, row, cols, col_map, agg['prior'], table_type=ctype)
        prior_row = row
        row += 1

        # Current week row
        ws.cell(row=row, column=2, value=fmt_date(agg['current']['week'])).border = border
        ws.cell(row=row, column=3, value='').border = border
        _write_data_row(ws, row, cols, col_map, agg['current'], table_type=ctype)
        curr_row = row
        ws.merge_cells(start_row=prior_row, start_column=3, end_row=curr_row, end_column=3)
        row += 1

        # % Change row — computed as real values, not formulas, to avoid
        # Excel caching issues with openpyxl-generated workbooks
        ws.cell(row=row, column=2, value="% Change").border = border
        ws.cell(row=row, column=2).fill = pctfill
        ws.cell(row=row, column=3, value='').border = border
        ws.cell(row=row, column=3).fill = pctfill

        def _pct_change(prior_val, curr_val):
            try:
                p = float(prior_val) if prior_val not in (None, '') else 0.0
                c_ = float(curr_val) if curr_val not in (None, '') else 0.0
                return round((c_ - p) / p, 6) if p != 0 else 0.0
            except (TypeError, ValueError):
                return 0.0

        for i, (cname, key) in enumerate(cols[2:], start=4):
            c = ws.cell(row=row, column=i)
            c.border = border
            c.fill = pctfill
            if key is None or key == 'suppress_blank':
                c.value = ''
            else:
                prior_cell = ws.cell(row=prior_row, column=i)
                curr_cell  = ws.cell(row=curr_row,  column=i)
                c.value = _pct_change(prior_cell.value, curr_cell.value)
                c.number_format = '0.0%'
        row += 2

    # Column widths
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    for i in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, skipped

# =============================================================================
# STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(page_title="WoW Report Generator", page_icon="📊", layout="wide")
    st.title("WoW Performance Update Report")
    st.markdown("Upload an SA360 export (CSV or Excel) to generate the weekly report.")

    uploaded = st.file_uploader(
        "Drop your file here",
        type=["csv", "xlsx", "xls"],
        help="Accepts SA360 CSV or Excel exports"
    )

    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    # ---- Load & Parse ----
    with st.spinner("Reading file..."):
        df, error = load_file(uploaded)

    if error:
        st.error(f"**Could not parse file:** {error}")
        st.markdown("**Tips:**")
        st.markdown("- Make sure the file is an SA360 export with columns like `Campaign`, `Week (Mon to Sun)`, `Cost`, `Clicks`, `Impr.`")
        st.markdown("- If column names have changed, let the team lead know so aliases can be updated.")
        return

    st.success(f"Loaded **{len(df):,}** rows from `{uploaded.name}`")

    # ---- Clean & Classify ----
    df = clean_numerics(df)
    df = add_classifications(df)
    weeks = sorted(df['Week (Mon to Sun)'].dropna().unique())

    # ---- Summary ----
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Weeks in file", len(weeks))
    with col2:
        st.metric("Total rows", f"{len(df):,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'index': 'Type', 'Customer Type': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c2:
            st.markdown("**Brand / NB**")
            st.dataframe(df['Brand/NB'].value_counts().reset_index().rename(
                columns={'index': 'Type', 'Brand/NB': 'Type', 'count': 'Rows'}
            ), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    # ---- Metrics Summary ----
    st.divider()
    st.subheader("📊 Overall Metrics by Week")
    st.caption("Total across all campaigns in the export — use this to sanity check scope before generating.")

    summary_metrics = ['Impr.', 'Clicks', 'Cost', 'Address Capture', 'Total Conversions - VBB']
    week_summary = []
    for week in weeks:
        wdata = df[df['Week (Mon to Sun)'] == week]
        row_data = {'Week': pd.to_datetime(week).strftime('%m/%d/%y') if not pd.isna(week) else str(week)}
        for m in summary_metrics:
            if m in wdata.columns:
                row_data[m] = wdata[m].sum()
        week_summary.append(row_data)

    summary_df = pd.DataFrame(week_summary)
    if 'Cost' in summary_df.columns:
        summary_df['Cost'] = summary_df['Cost'].apply(lambda x: f"${x:,.2f}")
    if 'Impr.' in summary_df.columns:
        summary_df['Impr.'] = summary_df['Impr.'].apply(lambda x: f"{int(x):,}")
    if 'Clicks' in summary_df.columns:
        summary_df['Clicks'] = summary_df['Clicks'].apply(lambda x: f"{int(x):,}")
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # ---- NC Non-Testing Diagnostic ----
    st.divider()
    st.subheader("🔍 NC Non-Testing Scope Check")
    st.caption("Shows what's being included vs excluded — NC Non-Testing should equal All NC minus the labeled buckets.")

    labels_col = 'Labels on Campaign: Directly Applied'
    account_col = next((c for c in df.columns if 'account' in c.lower()), None)
    latest_week = df[df['Week (Mon to Sun)'] == weeks[-1]]

    # All NC campaigns
    nc_all = latest_week[latest_week['Customer Type'] == 'NC']

    # Apply same exclusions as the report
    nc_testing = nc_all.copy()
    for label in NC_SPECIFIC_LABELS:
        if labels_col in nc_testing.columns:
            nc_testing = nc_testing[~nc_testing[labels_col].astype(str).str.contains(label, na=False)]
    for pattern in NC_EXCLUDE_CAMPAIGN_PATTERNS:
        nc_testing = nc_testing[~nc_testing['Campaign'].astype(str).str.contains(pattern, case=False, na=False)]
    if account_col:
        for acct in NC_EXCLUDE_ACCOUNTS:
            nc_testing = nc_testing[~nc_testing[account_col].astype(str).str.contains(acct, case=False, na=False)]

    excluded = nc_all[~nc_all.index.isin(nc_testing.index)]

    diag_cols = st.columns(3)
    with diag_cols[0]:
        st.metric("All NC campaigns (latest week)", f"{nc_all['Campaign'].nunique():,} campaigns")
        if 'Clicks' in nc_all.columns:
            st.metric("NC Clicks", f"{int(nc_all['Clicks'].sum()):,}")
        if 'Cost' in nc_all.columns:
            st.metric("NC Cost", f"${nc_all['Cost'].sum():,.2f}")
    with diag_cols[1]:
        st.metric("NC Non-Testing (after exclusions)", f"{nc_testing['Campaign'].nunique():,} campaigns")
        if 'Clicks' in nc_testing.columns:
            st.metric("Non-Testing Clicks", f"{int(nc_testing['Clicks'].sum()):,}")
        if 'Cost' in nc_testing.columns:
            st.metric("Non-Testing Cost", f"${nc_testing['Cost'].sum():,.2f}")
    with diag_cols[2]:
        st.metric("Excluded campaigns", f"{excluded['Campaign'].nunique():,} campaigns")
        if 'Clicks' in excluded.columns:
            st.metric("Excluded Clicks", f"{int(excluded['Clicks'].sum()):,}")
        if 'Cost' in excluded.columns:
            st.metric("Excluded Cost", f"${excluded['Cost'].sum():,.2f}")

    with st.expander("See excluded campaigns", expanded=False):
        if not excluded.empty and labels_col in excluded.columns:
            excl_display = excluded[['Campaign', labels_col] + ([account_col] if account_col else [])].drop_duplicates('Campaign')
            st.dataframe(excl_display, use_container_width=True, hide_index=True)
        else:
            st.write("No campaigns excluded or labels column not found.")

    with st.expander("See NC Non-Testing campaigns included", expanded=False):
        if not nc_testing.empty:
            incl_display = nc_testing[['Campaign'] + ([labels_col] if labels_col in nc_testing.columns else [])].drop_duplicates('Campaign')
            st.dataframe(incl_display, use_container_width=True, hide_index=True)


    # ---- Generate Report ----
    st.divider()

    if len(weeks) < 2:
        st.warning(f"Need at least 2 weeks of data to generate a WoW report. Found {len(weeks)} week(s).")
        return

    curr_week = pd.to_datetime(weeks[-1]).strftime('%Y-%m-%d')
    filename = f"WoW_Performance_Update_{curr_week}.xlsx"

    if st.button("Generate Report", type="primary", use_container_width=True):
        with st.spinner("Building Excel report..."):
            buf, skipped = create_report(df)

        if skipped:
            with st.expander(f"{len(skipped)} table(s) skipped — label mismatch or no data found", expanded=True):
                for name, reason in skipped:
                    st.markdown(f"- **{name}**: {reason}")
                st.markdown("Check the `TODO` comments in the `TABLES` config at the top of the script and verify the label strings match what's in your SA360 export.")

        st.download_button(
            label=f"Download {filename}",
            data=buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        st.success("Report ready! Click above to download.")

if __name__ == "__main__":
    main()
