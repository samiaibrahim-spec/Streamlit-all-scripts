#!/usr/bin/env python3
"""
Keyword Analysis Tool — Streamlit App
Analyzes top 10 keywords by Customer Type (CC/NC) and Brand/NB.

Usage: streamlit run keyword_analysis_app.py
"""

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import os
from datetime import datetime

# =============================================================================
# CAMPAIGN NAME PARSING
# =============================================================================

def parse_campaign_name(Campaign):
    """
    Parse campaign name to extract customer type, engine, and brand.
    Parts: [0]Brand [1]Channel [2]Type [3]CustomerType [4-6] [last]Engine

    Returns: dict with 'customer_type', 'engine', 'brand'
    """
    parts = Campaign.split('_')

    result = {
        'customer_type': None,
        'engine': None,
        'brand': None
    }

    campaign_lower = Campaign.lower()
    if 'google' in campaign_lower:
        result['engine'] = 'Google'
    elif 'bing' in campaign_lower:
        result['engine'] = 'Bing'

    # Customer Type: Look for CC or NC (typically 4th position)
    for part in parts:
        part_upper = part.upper()
        if part_upper in ['CC', 'NC']:
            result['customer_type'] = part_upper
            break

    # Brand: Look for Brand, Br, Nonbr, or NonBrand patterns
    for i, part in enumerate(parts):
        part_lower = part.lower()
        if part_lower == 'nonbr' or part_lower == 'nonbrand':
            result['brand'] = 'NonBrand'
            break
        elif part_lower == 'br' or part_lower == 'brand':
            result['brand'] = 'Brand'
            break

    return result


def add_parsed_columns(df, campaign_col='Campaign'):
    """
    Add Customer Type, Engine, and Brand columns by parsing the campaign name.
    """
    df = df.copy()

    parsed = df[campaign_col].apply(parse_campaign_name)

    df['Customer Type'] = parsed.apply(lambda x: x['customer_type'])
    df['Engine'] = parsed.apply(lambda x: x['engine'])
    df['Brand'] = parsed.apply(lambda x: x['brand'])

    unparsed_ctype = df[df['Customer Type'].isna()]
    if len(unparsed_ctype) > 0:
        st.warning(f"Could not parse Customer Type from {len(unparsed_ctype)} campaigns")

    unparsed_engine = df[df['Engine'].isna()]
    if len(unparsed_engine) > 0:
        st.warning(f"Could not parse Engine from {len(unparsed_engine)} campaigns")

    unparsed_brand = df[df['Brand'].isna()]
    if len(unparsed_brand) > 0:
        st.warning(f"Could not parse Brand from {len(unparsed_brand)} campaigns")

    return df

# =============================================================================
# FILE LOADING
# =============================================================================

def load_file(uploaded_file):
    """Load CSV or Excel file and return dataframe."""
    name = uploaded_file.name.lower()

    if name.endswith('.csv'):
        for enc in ("utf-8", "utf-8-sig", "utf-16"):
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, encoding=enc, sep=None, engine='python')
                if 'Campaign' in df.columns and 'Search keyword' in df.columns:
                    return df, None
            except (UnicodeDecodeError, Exception):
                continue

        # Try with skiprows
        for enc in ("utf-8", "utf-8-sig", "utf-16"):
            for skip in [1, 2, 3]:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=enc, sep=None, engine='python', skiprows=skip)
                    if 'Campaign' in df.columns and 'Search keyword' in df.columns:
                        return df, None
                except (UnicodeDecodeError, Exception):
                    continue

        return None, "Could not parse CSV. Ensure it contains 'Campaign' and 'Search keyword' columns."

    elif name.endswith(('.xlsx', '.xls')):
        for skip in [0, 1, 2, 3]:
            try:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, skiprows=skip)
                if 'Campaign' in df.columns and 'Search keyword' in df.columns:
                    return df, None
            except Exception:
                continue

        return None, "Could not parse Excel file. Ensure it contains 'Campaign' and 'Search keyword' columns."

    return None, f"Unsupported file type: {name}"

# =============================================================================
# KEYWORD ANALYSIS
# =============================================================================

def analyze_keywords(df, sort_by='Clicks'):
    """
    Analyze top 10 keywords by Customer Type (CC/NC) and Brand/NB.

    Returns:
        results: dict of segment results
        messages: list of log messages
    """
    messages = []

    if 'Campaign' in df.columns:
        df = add_parsed_columns(df, campaign_col='Campaign')

    messages.append(f"Loaded {len(df)} rows")

    # Validate required columns
    required = ['Search keyword', 'Customer Type', 'Brand']
    missing = [col for col in required if col not in df.columns]
    if missing:
        return None, [f"Missing required columns: {missing}"]

    # Metrics to aggregate
    metrics = ['Clicks', 'Cost', 'Visits']
    agg_metrics = [m for m in metrics if m in df.columns]

    # Convert metric columns to numeric
    for col in agg_metrics:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

    # Group by Customer Type, Brand/NB, and Search keyword
    grouped = df.groupby(['Customer Type', 'Brand', 'Search keyword'], as_index=False).agg(
        {m: 'sum' for m in agg_metrics}
    )

    # Get unique combinations
    combos = grouped[['Customer Type', 'Brand']].drop_duplicates().sort_values(
        ['Customer Type', 'Brand']
    ).values

    results = {}

    for cust_type, brand_type in combos:
        subset = grouped[
            (grouped['Customer Type'] == cust_type) &
            (grouped['Brand'] == brand_type)
        ].copy()

        sort_col = sort_by if sort_by in subset.columns else 'Clicks'
        top_10 = subset.nlargest(10, sort_col)

        display_cols = ['Search keyword', 'Clicks', 'Cost']
        available_display = [c for c in display_cols if c in top_10.columns]

        top_10_display = top_10[available_display].reset_index(drop=True)
        top_10_display.index = top_10_display.index + 1

        key = f"{cust_type}_{brand_type}".replace(' ', '')
        results[key] = {
            'data': top_10_display,
            'customer_type': cust_type,
            'brand_type': brand_type,
            'sort_metric': sort_col,
            'total_keywords': len(subset)
        }

        messages.append(
            f"{cust_type} - {brand_type}: {len(subset)} total keywords, top 10 by {sort_col}"
        )

    return results, messages

# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def create_excel_output(results):
    """Create formatted Excel workbook and return as BytesIO buffer."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for key, result in sorted(results.items()):
        ws = wb.create_sheet(key[:31])

        # Add title
        ws['A1'] = f"Top 10 Keywords: {result['customer_type']} - {result['brand_type']}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

        # Add metadata
        ws['A2'] = f"Total Keywords in Segment: {result['total_keywords']}"
        ws['A2'].font = Font(italic=True, size=10)

        # Add headers (rank + data columns)
        df = result['data'].reset_index()
        df.columns = ['Rank'] + list(result['data'].columns)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Add data rows
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=False)

                if col_idx > 2:
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        if len(df.columns) > 7:
            ws.column_dimensions['H'].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# =============================================================================
# STREAMLIT APP
# =============================================================================

def run():
    st.title("Keyword Analysis Tool")
    st.markdown("Upload an SA360 keyword export (CSV or Excel) to generate top 10 keyword tables by Customer Type and Brand/NB.")

    uploaded = st.file_uploader(
        "Drop your file here",
        type=["csv", "xlsx", "xls"],
        help="Accepts SA360 CSV or Excel exports with Campaign and Search keyword columns"
    )

    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    # ---- Load ----
    with st.spinner("Reading file..."):
        df, error = load_file(uploaded)

    if error:
        st.error(f"Could not parse file: {error}")
        return

    st.success(f"Loaded {len(df):,} rows from {uploaded.name}")

    # ---- Parse ----
    with st.spinner("Parsing campaign names..."):
        if 'Campaign' in df.columns:
            df = add_parsed_columns(df, campaign_col='Campaign')

    # ---- Summary info ----
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total rows", f"{len(df):,}")
    with col2:
        st.metric("Unique keywords", f"{df['Search keyword'].nunique():,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'Customer Type': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c2:
            st.markdown("**Engine**")
            st.dataframe(df['Engine'].value_counts().reset_index().rename(
                columns={'Engine': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c3:
            st.markdown("**Brand**")
            st.dataframe(df['Brand'].value_counts().reset_index().rename(
                columns={'Brand': 'Type', 'count': 'Rows'}
            ), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    # ---- Sort option ----
    st.divider()

    available_metrics = [m for m in ['Clicks', 'Cost', 'Visits'] if m in df.columns]
    sort_by = st.selectbox("Sort top keywords by:", available_metrics, index=0)

    # ---- Generate ----
    if st.button("Generate Analysis", type="primary", use_container_width=True):

        with st.spinner("Analyzing keywords..."):
            results, messages = analyze_keywords(df, sort_by=sort_by)

        if results is None:
            for msg in messages:
                st.error(msg)
            return

        # Show results inline
        for key, result in sorted(results.items()):
            with st.expander(
                f"{result['customer_type']} - {result['brand_type']}  "
                f"({result['total_keywords']} keywords, top 10 by {result['sort_metric']})",
                expanded=True
            ):
                st.dataframe(result['data'], use_container_width=True, hide_index=False)

        # Excel download
        st.markdown("---")
        with st.spinner("Building Excel file..."):
            excel_buf = create_excel_output(results)

        st.download_button(
            label="Download keyword_analysis_results.xlsx",
            data=excel_buf,
            file_name="keyword_analysis_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        st.success(f"Done. {len(results)} segment(s) analyzed.")


def main():
    st.set_page_config(page_title="Keyword Analysis Tool", layout="wide")
    run()


if __name__ == "__main__":
    main()
