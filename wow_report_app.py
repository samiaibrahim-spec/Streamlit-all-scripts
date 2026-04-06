#!/usr/bin/env python3
"""
WoW Performance Update Report Generator — Streamlit App
Converts SA360 CSV/XLSX export into formatted Excel report.

Usage: streamlit run wow_report_app.py
"""

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import os

# =============================================================================
# CONFIGURATION
# =============================================================================

METRICS = [
    'Impr.', 'Clicks', 'Cost', 'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New', 'Address Capture', 'Begin Checkout',
    'Main Sales Number', 'Contact Us Page', 'Quality Sales Call - AN',
    'Total Conversions - VBB', 'Total Conversion Value - VBB',
    'Chat Initiation - Order Services',
]

TOTAL_ACTIONS_COMPONENTS = [
    'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New',
    'Chat Initiation - Order Services',
    'Quality Sales Call - AN',
]


TABLES = [
    ('All SEM', {}, 'standard'),
    ('Brand SEM', {'Brand/NB': 'Brand'}, 'standard'),
    ('Nonbrand SEM', {'Brand/NB': 'NB'}, 'standard'),
    #('NC VBB Campaigns', {'Labels on Campaign: Directly Applied': '2026 VBB Google Campaigns', 'Customer Type': 'NC'}),
    #('NC VBB Campaigns', {'Test Segment': 'VBB', 'Customer Type': 'NC'}, 'vbb'),
    ('NC VBB Campaigns', {'Labels on Campaign: Directly Applied': '2026 VBB Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Internet Campaigns', {'Labels on Campaign: Directly Applied': 'CBB NB Internet Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC CBB NB Internet Campaigns', {'Test Segment': 'NB Internet CBB', 'Customer Type': 'NC'}, 'vbb'),
    ('NC UpMarket Campaigns', {'Labels on Campaign: Directly Applied': '2026 UpMarket Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC CBB NB Campaigns', {'Labels on Campaign: Directly Applied': 'NB CBB', 'Customer Type': 'NC'}, 'vbb'),
    ('NB Consolidated Campaigns', {'Labels on Campaign: Directly Applied': 'Nonbrand Consolidation 3.19.26'}, 'standard'),
    #('MSFT CBB NB Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT CBB NB Campaigns Feb 26', 'Customer Type': 'NC'}, 'standard'),
    ('MSFT CBB NB Campaigns', {'Test Segment': 'NB MSFT CBB', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Google Campaigns', {'Labels on Campaign: Directly Applied': '2026 CBB NB Remaining Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC CBB NB Google Campaigns', {'Test Segment': '2026 CBB NB Remaining Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC CBB MSFT NB Converting Campaigns', {'Labels on Campaign: Directly Applied': 'CB - NC - CBB MSFT NB Converting Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC CBB MSFT NB Converting Campaigns', {'Test Segment': 'CB - NC - CBB MSFT NB Converting Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC Max Clicks NB MSFT Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT NB Max Clicks Campaigns', 'Customer Type': 'NC'},'Standard'),
    ('NC Max Clicks NB MSFT Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT NB Max Clicks Campaigns', 'Customer Type': 'NC'},'vbb'),
    #('NC Max Clicks NB MSFT Campaigns', {'Test Segment': '2026 CBB NB Remaining MSFT Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    #('NC Non-Testing Campaigns', {'Test Segment': 'Non-Testing', 'Customer Type': 'NC'}, 'vbb'),
   ('NC Non-Testing Campaigns', {'Labels on Campaign: Directly Applied': 'Current NC Non-Testing', 'Customer Type': 'NC'}, 'vbb')
]

STANDARD_COLS = [
    ('Date Range', None), ('Tactic', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Main Sales Number', 'Main Sales Number'), ('Contact Us Page', 'Contact Us Page'),
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Actions', 'total_actions'),
    ('Cost per Action', 'cpactions'),
]

VBB_COLS = [
    ('Date Range', None), ('Campaign', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Conversions - VBB', 'Total Conversions - VBB'),
    ('Total Conversion Value - VBB', 'Total Conversion Value - VBB'),
    ('Total Actions', 'total_actions'),
    ('Cost per Action', 'cost per action'),
]

# Column aliases — maps the standard name to known alternatives
COLUMN_ALIASES = {
    'Campaign': ['Campaign', 'Campaign Name', 'campaign'],
    'Week (Mon to Sun)': ['Week (Mon to Sun)', 'Week', 'week'],
    'Cost': ['Cost', 'Spend', 'cost', 'spend'],
    'Clicks': ['Clicks', 'clicks'],
    'Impr.': ['Impr.', 'Impressions', 'Impr', 'impressions', 'impr.'],
    'Labels on Campaign: Directly Applied': [
        'Labels on Campaign: Directly Applied',
        'Labels on campaign: Directly applied',
        'Labels',
        'Campaign Labels',
    ],
}

# =============================================================================
# CLASSIFICATION FUNCTIONS
# =============================================================================

def classify_customer_type(campaign):
    if pd.isna(campaign):
        return 'NC'
    return 'CC' if '_CC_' in campaign else 'NC'


def classify_brand_nb(campaign):
    if pd.isna(campaign):
        return 'Brand'
    if '_Nonbr_' in campaign:
        return 'NB'
    return 'Brand'


def add_classifications(df):
    df = df.copy()
    df['Customer Type'] = df['Campaign'].apply(classify_customer_type)
    df['Brand/NB'] = df['Campaign'].apply(classify_brand_nb)

    labels_col = 'Labels on Campaign: Directly Applied'
    if labels_col not in df.columns:
        df[labels_col] = ''

    return df

# =============================================================================
# DATA LOADING AND PROCESSING
# =============================================================================

def normalize_columns(df):
    """Rename columns to standard names using COLUMN_ALIASES."""
    df.columns = df.columns.str.strip()
    rename_map = {}
    for standard_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns and alias != standard_name:
                rename_map[alias] = standard_name
                break
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def check_required_columns(df):
    """Check if required columns are present. Returns (ok, missing_list)."""
    required = ['Campaign', 'Week (Mon to Sun)', 'Cost', 'Clicks', 'Impr.']
    missing = [c for c in required if c not in df.columns]
    return len(missing) == 0, missing


def load_file(uploaded_file):
    """Load CSV or Excel file, normalize columns, and validate."""
    name = uploaded_file.name.lower()
    errors = []

    # --- Excel ---
    if name.endswith(('.xlsx', '.xls')):
        for skip in [2, 0, 1, 3]:
            try:
                df = pd.read_excel(uploaded_file, skiprows=skip)
                df = normalize_columns(df)
                ok, missing = check_required_columns(df)
                if ok:
                    return df, None
            except Exception:
                continue
            finally:
                uploaded_file.seek(0)

        # Last attempt — show what we found
        df = pd.read_excel(uploaded_file, skiprows=2)
        uploaded_file.seek(0)
        df = normalize_columns(df)
        _, missing = check_required_columns(df)
        return None, f"Could not find required columns in Excel file. Missing: {missing}. Found: {list(df.columns[:15])}"

    # --- CSV / TSV ---
    attempts = [
        ('utf-16', '\t', 2), ('utf-16-le', '\t', 2),
        ('utf-8', ',', 2), ('utf-8', '\t', 2), ('utf-8', ';', 2),
        ('utf-8', ',', 0), ('utf-8', '\t', 0), ('utf-8', ';', 0),
        ('latin-1', ',', 2), ('latin-1', '\t', 2), ('latin-1', ';', 2),
        ('latin-1', ',', 0), ('latin-1', '\t', 0), ('latin-1', ';', 0),
    ]

    last_columns = []
    for enc, sep, skip in attempts:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=enc, sep=sep, skiprows=skip)
            df = normalize_columns(df)
            ok, missing = check_required_columns(df)
            if ok:
                return df, None
            last_columns = list(df.columns[:15])
        except Exception:
            continue

    return None, f"Could not parse CSV. Last columns found: {last_columns}"


def clean_numerics(df):
    for col in METRICS:
        if col in df.columns:
            df[col] = df[col].replace(['--', ' --', '- ', '-'], 0)
            df[col] = df[col].replace(r'[\$,US]', '', regex=True)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


def aggregate(df, filters):
    filt = df.copy()
    for col, val in filters.items():
        if col in filt.columns:
            if col == 'Labels on Campaign: Directly Applied':
                filt = filt[filt[col].astype(str).str.contains(val, na=False)]
            else:
                filt = filt[filt[col] == val]

    weeks = sorted(filt['Week (Mon to Sun)'].dropna().unique())
    if len(weeks) < 2:
        raise ValueError(f"Need 2+ weeks, found {len(weeks)}")

    result = {}
    for label, week in [('current', weeks[-1]), ('prior', weeks[-2])]:
        wdata = filt[filt['Week (Mon to Sun)'] == week]
        agg = {c: wdata[c].sum() if c in wdata.columns else 0 for c in METRICS}
        agg['week'] = week
        result[label] = agg
    return result

# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def fmt_date(week):
    if pd.isna(week):
        return ''
    try:
        s = pd.to_datetime(week)
        e = s + pd.Timedelta(days=6)
        return f"{s.month}/{s.day}-{e.month}/{e.day}"
    except Exception:
        return str(week)


def _build_column_map(cols):
    col_map = {}
    for i, (cname, key) in enumerate(cols, start=2):
        if key is not None:
            col_map[key] = get_column_letter(i)
    return col_map


def _write_data_row(ws, row, cols, col_map, agg_data):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, (cname, key) in enumerate(cols[2:], start=4):
        c = ws.cell(row=row, column=i)
        c.border = border
        if key == 'cpc':
            c.value = f"=IF(E{row}=0,0,F{row}/E{row})"
            c.number_format = '$#,##0.00'
        elif key == 'ctr':
            c.value = f"=IF(D{row}=0,0,E{row}/D{row})"
            c.number_format = '0.00%'
        elif key == 'total_actions':
            refs = [f"{col_map[k]}{row}" for k in TOTAL_ACTIONS_COMPONENTS if k in col_map]
            c.value = f"={'+'.join(refs)}" if refs else 0
            c.number_format = '#,##0'
        elif key == 'cpactions':
            cost_col = col_map.get('Cost', 'F')
            ta_col = col_map.get('total_actions', '')
            c.value = f"=IF({ta_col}{row}=0,0,{cost_col}{row}/{ta_col}{row})" if ta_col else 0
            c.number_format = '$#,##0.00'
        elif key is None:
            c.value = ''
        else:
            v = agg_data.get(key, 0)
            c.value = int(v) if isinstance(v, float) and v == int(v) else v
            if 'Cost' in cname or 'Value' in cname:
                c.number_format = '$#,##0.00'


def create_report(df):
    """Generate the Excel report and return it as a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WoW Performance Update"

    hfont = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    pctfill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    valign = Alignment(vertical='center')

    row = 1
    skipped = []

    for tname, filters, ctype in TABLES:
        try:
            agg = aggregate(df, filters)
        except Exception as e:
            skipped.append((tname, str(e)))
            continue

        cols = STANDARD_COLS if ctype == 'standard' else VBB_COLS
        col_map = _build_column_map(cols)

        # Header row
        for i, (cname, _) in enumerate(cols, start=2):
            c = ws.cell(row=row, column=i, value=cname)
            c.font = hfont
            c.border = border
        row += 1

        # Prior week row
        ws.cell(row=row, column=2, value=fmt_date(agg['prior']['week'])).border = border
        tactic_cell = ws.cell(row=row, column=3, value=tname)
        tactic_cell.border = border
        tactic_cell.alignment = valign
        _write_data_row(ws, row, cols, col_map, agg['prior'])
        prior_row = row
        row += 1

        # Current week row
        ws.cell(row=row, column=2, value=fmt_date(agg['current']['week'])).border = border
        ws.cell(row=row, column=3, value='').border = border
        _write_data_row(ws, row, cols, col_map, agg['current'])
        curr_row = row
        ws.merge_cells(start_row=prior_row, start_column=3, end_row=curr_row, end_column=3)
        row += 1

        # % Change row
        ws.cell(row=row, column=2, value="% Change").border = border
        ws.cell(row=row, column=2).fill = pctfill
        ws.cell(row=row, column=3, value='').border = border
        ws.cell(row=row, column=3).fill = pctfill

        for i, (cname, key) in enumerate(cols[2:], start=4):
            c = ws.cell(row=row, column=i)
            c.border = border
            c.fill = pctfill
            if key is None:
                c.value = ''
            else:
                L = get_column_letter(i)
                c.value = f"=IF({L}{prior_row}=0,0,({L}{curr_row}-{L}{prior_row})/{L}{prior_row})"
                c.number_format = '0.0%'

        row += 2

    # Column widths
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    for i in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, skipped

# =============================================================================
# STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(page_title="WoW Report Generator", page_icon="📊", layout="wide")

    st.title(" WoW Performance Update Report")
    st.markdown("Upload an SA360 export (CSV or Excel) to generate the weekly report.")

    uploaded = st.file_uploader(
        "Drop your file here",
        type=["csv", "xlsx", "xls"],
        help="Accepts SA360 CSV or Excel exports"
    )

    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    # ---- Load & Parse ----
    with st.spinner("Reading file..."):
        df, error = load_file(uploaded)

    if error:
        st.error(f"**Could not parse file:** {error}")
        st.markdown("**Tips:**")
        st.markdown("- Make sure the file is an SA360 export with columns like `Campaign`, `Week (Mon to Sun)`, `Cost`, `Clicks`, `Impr.`")
        st.markdown("- If column names have changed, let the team lead know so aliases can be updated.")
        return

    st.success(f"Loaded **{len(df):,}** rows from `{uploaded.name}`")

    # ---- Clean & Classify ----
    df = clean_numerics(df)
    df = add_classifications(df)

    weeks = sorted(df['Week (Mon to Sun)'].dropna().unique())

    # ---- Summary ----
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Weeks in file", len(weeks))
    with col2:
        st.metric("Total rows", f"{len(df):,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'index': 'Type', 'Customer Type': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c2:
            st.markdown("**Brand / NB**")
            st.dataframe(df['Brand/NB'].value_counts().reset_index().rename(
                columns={'index': 'Type', 'Brand/NB': 'Type', 'count': 'Rows'}
            ), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    # ---- Generate Report ----
    st.divider()

    if len(weeks) < 2:
        st.warning(f"Need at least 2 weeks of data to generate a WoW report. Found {len(weeks)} week(s).")
        return

    curr_week = pd.to_datetime(weeks[-1]).strftime('%Y-%m-%d')
    filename = f"WoW_Performance_Update_{curr_week}.xlsx"

    if st.button(" Generate Report", type="primary", use_container_width=True):
        with st.spinner("Building Excel report..."):
            buf, skipped = create_report(df)

        if skipped:
            with st.expander(f" {len(skipped)} table(s) skipped", expanded=False):
                for name, reason in skipped:
                    st.markdown(f"- **{name}**: {reason}")

        st.download_button(
            label=f" Download {filename}",
            data=buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        st.success("Report ready! Click above to download.")


if __name__ == "__main__":
    main()
