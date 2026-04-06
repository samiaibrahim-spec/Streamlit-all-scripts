#!/usr/bin/env python3
"""
CB Reporting Hub — Streamlit App
Combines Weekly WoW Report, Monthly Campaign Summary, and Keyword Analysis
into a single app with a report selector.

Usage: streamlit run app.py
"""

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
import os


# =============================================================================
#
#   WEEKLY WOW REPORT
#
# =============================================================================

# ---- Configuration ----

WEEKLY_METRICS = [
    'Impr.', 'Clicks', 'Cost', 'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New', 'Address Capture', 'Begin Checkout',
    'Main Sales Number', 'Contact Us Page', 'Quality Sales Call - AN',
    'Total Conversions - VBB', 'Total Conversion Value - VBB',
    'Chat Initiation - Order Services',
]

TOTAL_ACTIONS_COMPONENTS = [
    'CB eCom Order Tag - New',
    'CB General Lead Form Submission - New',
    'Chat Initiation - Order Services',
    'Quality Sales Call - AN',
]

WEEKLY_TABLES = [
    ('All SEM', {}, 'standard'),
    ('Brand SEM', {'Brand/NB': 'Brand'}, 'standard'),
    ('Nonbrand SEM', {'Brand/NB': 'NB'}, 'standard'),
    ('NC VBB Campaigns', {'Labels on Campaign: Directly Applied': '2026 VBB Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Internet Campaigns', {'Labels on Campaign: Directly Applied': 'CBB NB Internet Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC UpMarket Campaigns', {'Labels on Campaign: Directly Applied': '2026 UpMarket Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Campaigns', {'Labels on Campaign: Directly Applied': 'NB CBB', 'Customer Type': 'NC'}, 'vbb'),
    ('MSFT CBB NB Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT CBB NB Campaigns Feb 26', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB NB Google Campaigns', {'Labels on Campaign: Directly Applied': '2026 CBB NB Remaining Google Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC CBB MSFT NB Converting Campaigns', {'Labels on Campaign: Directly Applied': 'CB - NC - CBB MSFT NB Converting Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC Max Clicks NB MSFT Campaigns', {'Labels on Campaign: Directly Applied': 'MSFT NB Max Clicks Campaigns', 'Customer Type': 'NC'}, 'vbb'),
    ('NC Non-Testing Campaigns', {'Labels on Campaign: Directly Applied': 'Current NC Non-Testing', 'Customer Type': 'NC'}, 'vbb'),
]

STANDARD_COLS = [
    ('Date Range', None), ('Tactic', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Main Sales Number', 'Main Sales Number'), ('Contact Us Page', 'Contact Us Page'),
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Actions', 'total_actions'),
    ('Cost per Action', 'cpactions'),
]

VBB_COLS = [
    ('Date Range', None), ('Campaign', None), ('Impr.', 'Impr.'), ('Clicks', 'Clicks'),
    ('Cost', 'Cost'), ('Avg. CPC', 'cpc'), ('Avg. CTR', 'ctr'),
    ('eCom Orders', 'CB eCom Order Tag - New'),
    ('Lead Form Submissions', 'CB General Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Quality Sales Calls', 'Quality Sales Call - AN'),
    ('Chat Initiation', 'Chat Initiation - Order Services'),
    ('Total Conversions - VBB', 'Total Conversions - VBB'),
    ('Total Conversion Value - VBB', 'Total Conversion Value - VBB'),
    ('Total Actions', 'total_actions'),
    ('Cost per Action', 'cpactions'),
]

WEEKLY_COLUMN_ALIASES = {
    'Campaign': ['Campaign', 'Campaign Name', 'campaign'],
    'Week (Mon to Sun)': ['Week (Mon to Sun)', 'Week', 'week'],
    'Cost': ['Cost', 'Spend', 'cost', 'spend'],
    'Clicks': ['Clicks', 'clicks'],
    'Impr.': ['Impr.', 'Impressions', 'Impr', 'impressions', 'impr.'],
    'Labels on Campaign: Directly Applied': [
        'Labels on Campaign: Directly Applied',
        'Labels on campaign: Directly applied',
        'Labels',
        'Campaign Labels',
    ],
}

# ---- Weekly classification ----

def weekly_classify_customer_type(campaign):
    if pd.isna(campaign):
        return 'NC'
    return 'CC' if '_CC_' in campaign else 'NC'


def weekly_classify_brand_nb(campaign):
    if pd.isna(campaign):
        return 'Brand'
    if '_Nonbr_' in campaign:
        return 'NB'
    return 'Brand'


def weekly_add_classifications(df):
    df = df.copy()
    df['Customer Type'] = df['Campaign'].apply(weekly_classify_customer_type)
    df['Brand/NB'] = df['Campaign'].apply(weekly_classify_brand_nb)
    labels_col = 'Labels on Campaign: Directly Applied'
    if labels_col not in df.columns:
        df[labels_col] = ''
    return df

# ---- Weekly file loading ----

def weekly_normalize_columns(df):
    df.columns = df.columns.str.strip()
    rename_map = {}
    for standard_name, aliases in WEEKLY_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns and alias != standard_name:
                rename_map[alias] = standard_name
                break
    if rename_map:
        df = df.rename(columns=rename_map)
    return df


def weekly_check_required(df):
    required = ['Campaign', 'Week (Mon to Sun)', 'Cost', 'Clicks', 'Impr.']
    missing = [c for c in required if c not in df.columns]
    return len(missing) == 0, missing


def weekly_load_file(uploaded_file):
    name = uploaded_file.name.lower()

    if name.endswith(('.xlsx', '.xls')):
        for skip in [2, 0, 1, 3]:
            try:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, skiprows=skip)
                df = weekly_normalize_columns(df)
                ok, missing = weekly_check_required(df)
                if ok:
                    return df, None
            except Exception:
                continue
            finally:
                uploaded_file.seek(0)
        return None, "Could not find required columns in Excel file."

    attempts = [
        ('utf-16', '\t', 2), ('utf-16-le', '\t', 2),
        ('utf-8', ',', 2), ('utf-8', '\t', 2), ('utf-8', ';', 2),
        ('utf-8', ',', 0), ('utf-8', '\t', 0), ('utf-8', ';', 0),
        ('latin-1', ',', 2), ('latin-1', '\t', 2), ('latin-1', ';', 2),
        ('latin-1', ',', 0), ('latin-1', '\t', 0), ('latin-1', ';', 0),
    ]
    last_columns = []
    for enc, sep, skip in attempts:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=enc, sep=sep, skiprows=skip)
            df = weekly_normalize_columns(df)
            ok, missing = weekly_check_required(df)
            if ok:
                return df, None
            last_columns = list(df.columns[:15])
        except Exception:
            continue
    return None, f"Could not parse CSV. Last columns found: {last_columns}"


def weekly_clean_numerics(df):
    for col in WEEKLY_METRICS:
        if col in df.columns:
            df[col] = df[col].replace(['--', ' --', '- ', '-'], 0)
            df[col] = df[col].replace(r'[\$,US]', '', regex=True)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df


def weekly_aggregate(df, filters):
    filt = df.copy()
    for col, val in filters.items():
        if col in filt.columns:
            if col == 'Labels on Campaign: Directly Applied':
                filt = filt[filt[col].astype(str).str.contains(val, na=False)]
            else:
                filt = filt[filt[col] == val]
    weeks = sorted(filt['Week (Mon to Sun)'].dropna().unique())
    if len(weeks) < 2:
        raise ValueError(f"Need 2+ weeks, found {len(weeks)}")
    result = {}
    for label, week in [('current', weeks[-1]), ('prior', weeks[-2])]:
        wdata = filt[filt['Week (Mon to Sun)'] == week]
        agg = {c: wdata[c].sum() if c in wdata.columns else 0 for c in WEEKLY_METRICS}
        agg['week'] = week
        result[label] = agg
    return result

# ---- Weekly Excel output ----

def weekly_fmt_date(week):
    if pd.isna(week):
        return ''
    try:
        s = pd.to_datetime(week)
        e = s + pd.Timedelta(days=6)
        return f"{s.month}/{s.day}-{e.month}/{e.day}"
    except Exception:
        return str(week)


def weekly_build_column_map(cols):
    col_map = {}
    for i, (cname, key) in enumerate(cols, start=2):
        if key is not None:
            col_map[key] = get_column_letter(i)
    return col_map


def weekly_write_data_row(ws, row, cols, col_map, agg_data):
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, (cname, key) in enumerate(cols[2:], start=4):
        c = ws.cell(row=row, column=i)
        c.border = border
        if key == 'cpc':
            c.value = f"=IF(E{row}=0,0,F{row}/E{row})"
            c.number_format = '$#,##0.00'
        elif key == 'ctr':
            c.value = f"=IF(D{row}=0,0,E{row}/D{row})"
            c.number_format = '0.00%'
        elif key == 'total_actions':
            refs = [f"{col_map[k]}{row}" for k in TOTAL_ACTIONS_COMPONENTS if k in col_map]
            c.value = f"={'+'.join(refs)}" if refs else 0
            c.number_format = '#,##0'
        elif key == 'cpactions':
            cost_col = col_map.get('Cost', 'F')
            ta_col = col_map.get('total_actions', '')
            c.value = f"=IF({ta_col}{row}=0,0,{cost_col}{row}/{ta_col}{row})" if ta_col else 0
            c.number_format = '$#,##0.00'
        elif key is None:
            c.value = ''
        else:
            v = agg_data.get(key, 0)
            c.value = int(v) if isinstance(v, float) and v == int(v) else v
            if 'Cost' in cname or 'Value' in cname:
                c.number_format = '$#,##0.00'


def weekly_create_report(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "WoW Performance Update"
    hfont = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    pctfill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    valign = Alignment(vertical='center')
    row = 1
    skipped = []
    for tname, filters, ctype in WEEKLY_TABLES:
        try:
            agg = weekly_aggregate(df, filters)
        except Exception as e:
            skipped.append((tname, str(e)))
            continue
        cols = STANDARD_COLS if ctype == 'standard' else VBB_COLS
        col_map = weekly_build_column_map(cols)
        for i, (cname, _) in enumerate(cols, start=2):
            c = ws.cell(row=row, column=i, value=cname)
            c.font = hfont
            c.border = border
        row += 1
        ws.cell(row=row, column=2, value=weekly_fmt_date(agg['prior']['week'])).border = border
        tactic_cell = ws.cell(row=row, column=3, value=tname)
        tactic_cell.border = border
        tactic_cell.alignment = valign
        weekly_write_data_row(ws, row, cols, col_map, agg['prior'])
        prior_row = row
        row += 1
        ws.cell(row=row, column=2, value=weekly_fmt_date(agg['current']['week'])).border = border
        ws.cell(row=row, column=3, value='').border = border
        weekly_write_data_row(ws, row, cols, col_map, agg['current'])
        curr_row = row
        ws.merge_cells(start_row=prior_row, start_column=3, end_row=curr_row, end_column=3)
        row += 1
        ws.cell(row=row, column=2, value="% Change").border = border
        ws.cell(row=row, column=2).fill = pctfill
        ws.cell(row=row, column=3, value='').border = border
        ws.cell(row=row, column=3).fill = pctfill
        for i, (cname, key) in enumerate(cols[2:], start=4):
            c = ws.cell(row=row, column=i)
            c.border = border
            c.fill = pctfill
            if key is None:
                c.value = ''
            else:
                L = get_column_letter(i)
                c.value = f"=IF({L}{prior_row}=0,0,({L}{curr_row}-{L}{prior_row})/{L}{prior_row})"
                c.number_format = '0.0%'
        row += 2
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    for i in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, skipped

# ---- Weekly UI ----

def run_weekly_report():
    st.header("Weekly WoW Performance Report")
    st.markdown("Upload an SA360 weekly export (CSV or Excel) to generate the WoW report.")

    uploaded = st.file_uploader("Drop your file here", type=["csv", "xlsx", "xls"], key="weekly_upload")
    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    with st.spinner("Reading file..."):
        df, error = weekly_load_file(uploaded)
    if error:
        st.error(f"Could not parse file: {error}")
        return

    st.success(f"Loaded {len(df):,} rows from {uploaded.name}")
    df = weekly_clean_numerics(df)
    df = weekly_add_classifications(df)
    weeks = sorted(df['Week (Mon to Sun)'].dropna().unique())

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Weeks in file", len(weeks))
    with col2:
        st.metric("Total rows", f"{len(df):,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'Customer Type': 'Type', 'count': 'Rows'}), hide_index=True)
        with c2:
            st.markdown("**Brand / NB**")
            st.dataframe(df['Brand/NB'].value_counts().reset_index().rename(
                columns={'Brand/NB': 'Type', 'count': 'Rows'}), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    st.divider()
    if len(weeks) < 2:
        st.warning(f"Need at least 2 weeks of data. Found {len(weeks)} week(s).")
        return

    curr_week = pd.to_datetime(weeks[-1]).strftime('%Y-%m-%d')
    filename = f"WoW_Performance_Update_{curr_week}.xlsx"

    if st.button("Generate Report", type="primary", use_container_width=True, key="weekly_btn"):
        with st.spinner("Building Excel report..."):
            buf, skipped = weekly_create_report(df)
        if skipped:
            with st.expander(f"{len(skipped)} table(s) skipped", expanded=False):
                for name, reason in skipped:
                    st.markdown(f"- **{name}**: {reason}")
        st.download_button(
            label=f"Download {filename}",
            data=buf, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
        st.success("Report ready.")


# =============================================================================
#
#   MONTHLY CAMPAIGN SUMMARY
#
# =============================================================================

# ---- Monthly parsing ----

def monthly_parse_campaign_name(campaign_name):
    parts = campaign_name.split('_')
    result = {'Customer_type': None, 'engine': None, 'brand': None}
    for part in parts:
        part_clean = part.strip().lower()
        if 'google' in part_clean:
            result['engine'] = 'Google'
            break
        elif 'bing' in part_clean:
            result['engine'] = 'Bing'
            break
    for part in parts:
        part_upper = part.upper()
        if part_upper in ['CC', 'NC']:
            result['Customer_type'] = part_upper
            break
    for i, part in enumerate(parts):
        part_lower = part.lower()
        if part_lower == 'nonbr' or part_lower == 'nonbrand':
            result['brand'] = 'NonBrand'
            break
        elif part_lower == 'br' or part_lower == 'brand':
            result['brand'] = 'Brand'
            break
    return result


def monthly_add_parsed_columns(df, campaign_col='Campaign'):
    df = df.copy()
    parsed = df[campaign_col].apply(monthly_parse_campaign_name)
    df['Customer Type'] = parsed.apply(lambda x: x['Customer_type'])
    df['Engine'] = parsed.apply(lambda x: x['engine'])
    df['Brand'] = parsed.apply(lambda x: x['brand'])
    unparsed_ctype = df[df['Customer Type'].isna()]
    if len(unparsed_ctype) > 0:
        st.warning(f"Could not parse Customer Type from {len(unparsed_ctype)} campaigns")
    unparsed_engine = df[df['Engine'].isna()]
    if len(unparsed_engine) > 0:
        st.warning(f"Could not parse Engine from {len(unparsed_engine)} campaigns")
    unparsed_brand = df[df['Brand'].isna()]
    if len(unparsed_brand) > 0:
        st.warning(f"Could not parse Brand from {len(unparsed_brand)} campaigns")
    return df

# ---- Monthly data prep ----

def monthly_prepare_dataframe(df):
    if pd.api.types.is_datetime64_any_dtype(df['Month']):
        df['Month_dt'] = pd.to_datetime(df['Month'])
    else:
        formats = [('%y-%b', 'YY-Mon'), ('%b-%y', 'Mon-YY'), ('%B %Y', 'Month YYYY')]
        parsed = False
        for fmt, desc in formats:
            attempt = pd.to_datetime(df['Month'], format=fmt, errors='coerce')
            if attempt.notna().sum() > 0:
                df['Month_dt'] = attempt
                parsed = True
                break
        if not parsed:
            df['Month_dt'] = pd.to_datetime(df['Month'], errors='coerce')
    return df

# ---- Monthly file loading ----

def monthly_load_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        for enc in ("utf-8-sig", "utf-8", "utf-16"):
            for skip in [2, 0]:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=enc, skiprows=skip)
                    if 'Campaign' in df.columns and 'Month' in df.columns:
                        return df, None
                except Exception:
                    continue
        return None, "Could not parse CSV. Ensure it contains 'Campaign' and 'Month' columns."
    elif name.endswith(('.xlsx', '.xls')):
        for skip in [0, 2, 1, 3]:
            try:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, skiprows=skip)
                if 'Campaign' in df.columns and 'Month' in df.columns:
                    return df, None
            except Exception:
                continue
        return None, "Could not parse Excel file. Ensure it contains 'Campaign' and 'Month' columns."
    return None, f"Unsupported file type: {name}"

# ---- Monthly summary tables ----

MONTHLY_SUMMARY_METRICS = [
    ('Spend', 'Cost'), ('Impressions', 'Impr.'), ('Clicks', 'Clicks'),
    ('avg CPC', None), ('avg CTR', None),
    ('eCom Orders', 'eCom Order - New'),
    ('Lead Form', 'Lead Form Submission - New'),
    ('Address Capture', 'Address Capture'), ('Begin Checkout', 'Begin Checkout'),
    ('Total Conversions - VBB', 'Total Conversions - VBB'),
    ('Chat Initiation - Order Services', 'Chat Initiation - Order Services'),
    ('Quality Sales Calls (Offline)', 'Quality Sales Call - AN'),
    ('Main Sales Number', 'Main Sales Number'), ('Contact Us Page', 'Contact Us Page')
]


def monthly_create_summary_table(df, customer_type, filter_col, filter_value,
                                  current_month_dt, prev_month_dt, year_ago_month_dt):
    df_filtered = df[df['Customer Type'] == customer_type].copy()
    if filter_col and filter_value:
        df_filtered = df_filtered[df_filtered[filter_col] == filter_value]
    table_name = f"{customer_type} - Overall" if filter_col is None else f"{customer_type} - {filter_value}"
    data = {table_name: []}
    for display_name, _ in MONTHLY_SUMMARY_METRICS:
        data[display_name] = []
    periods = [('current', current_month_dt), ('previous', prev_month_dt), ('year_ago', year_ago_month_dt)]
    for period_name, month_dt in periods:
        if month_dt is None:
            continue
        month_data = df_filtered[df_filtered['Month_dt'] == month_dt]
        if len(month_data) == 0:
            continue
        month_display = month_data['Month'].iloc[0]
        if isinstance(month_display, pd.Timestamp):
            month_display = month_display.strftime('%b-%y')
        data[table_name].append(str(month_display))
        for display_name, actual_col in MONTHLY_SUMMARY_METRICS:
            if display_name == 'avg CPC':
                total_cost = data['Spend'][-1] if data['Spend'] else 0
                total_clicks = data['Clicks'][-1] if data['Clicks'] else 0
                data[display_name].append(round(total_cost / total_clicks, 2) if total_clicks > 0 else 0)
            elif display_name == 'avg CTR':
                total_impr = data['Impressions'][-1] if data['Impressions'] else 0
                total_clicks = data['Clicks'][-1] if data['Clicks'] else 0
                data[display_name].append(round((total_clicks / total_impr * 100) if total_impr > 0 else 0, 2))
            elif actual_col in month_data.columns:
                value = month_data[actual_col].sum()
                try:
                    value = pd.to_numeric(value, errors='coerce')
                    data[display_name].append(round(value, 2) if pd.notna(value) else 0)
                except (ValueError, TypeError):
                    data[display_name].append(0)
            else:
                data[display_name].append(0)
    data[table_name].extend(['MoM', 'YoY'])
    for display_name, _ in MONTHLY_SUMMARY_METRICS:
        data[display_name].extend([None, None])
    return pd.DataFrame(data)


def monthly_create_formatted_summaries(df):
    summary_tables = {}
    if 'Month_dt' not in df.columns:
        df = monthly_prepare_dataframe(df)
    df_sorted = df.sort_values('Month_dt', ascending=False)
    unique_months = df_sorted.drop_duplicates(subset=['Month_dt'])[['Month', 'Month_dt']].reset_index(drop=True)
    if len(unique_months) < 2:
        return summary_tables
    current_month_dt = unique_months.iloc[0]['Month_dt']
    prev_month_dt = current_month_dt - pd.DateOffset(months=1)
    prev_match = df_sorted[df_sorted['Month_dt'] <= prev_month_dt].drop_duplicates(subset=['Month_dt'])
    if len(prev_match) > 0:
        prev_month_dt = prev_match.iloc[0]['Month_dt']
    elif len(unique_months) >= 2:
        prev_month_dt = unique_months.iloc[1]['Month_dt']
    else:
        prev_month_dt = None
    year_ago_dt = current_month_dt - pd.DateOffset(years=1)
    ya_matches = df_sorted[
        (df_sorted['Month_dt'] >= year_ago_dt - pd.Timedelta(days=45)) &
        (df_sorted['Month_dt'] <= year_ago_dt + pd.Timedelta(days=45))
    ].drop_duplicates(subset=['Month_dt'])
    if len(ya_matches) > 0:
        ya_matches = ya_matches.copy()
        ya_matches['diff'] = abs((ya_matches['Month_dt'] - year_ago_dt).dt.days)
        year_ago_month_dt = ya_matches.sort_values('diff').iloc[0]['Month_dt']
    else:
        year_ago_month_dt = None
    if 'Category (with Brand vs NB)' in df.columns and 'Campaign Type' not in df.columns:
        df['Campaign Type'] = df['Category (with Brand vs NB)'].apply(
            lambda x: 'Brand' if 'Brand' in str(x) and 'NB' not in str(x) else 'NonBrand')
    if 'Customer Type' not in df.columns:
        return summary_tables
    for cust_type in df['Customer Type'].dropna().unique():
        summary_tables[f'{cust_type} - Overall'] = monthly_create_summary_table(
            df, cust_type, None, None, current_month_dt, prev_month_dt, year_ago_month_dt)
        if 'Engine' in df.columns:
            for engine in df[df['Customer Type'] == cust_type]['Engine'].dropna().unique():
                summary_tables[f'{cust_type} - {engine}'] = monthly_create_summary_table(
                    df, cust_type, 'Engine', engine, current_month_dt, prev_month_dt, year_ago_month_dt)
        if 'Brand' in df.columns:
            for camp_type in df[df['Customer Type'] == cust_type]['Brand'].dropna().unique():
                summary_tables[f'{cust_type} - {camp_type}'] = monthly_create_summary_table(
                    df, cust_type, 'Brand', camp_type, current_month_dt, prev_month_dt, year_ago_month_dt)
    return summary_tables

# ---- Monthly Excel output ----

def monthly_format_summary_tables(workbook, sheet_names):
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    period_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    period_font = Font(bold=True)
    mom_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    yoy_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for sheet_name in sheet_names:
        sn = sheet_name[:31]
        if sn not in workbook.sheetnames:
            continue
        ws = workbook[sn]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        max_row = ws.max_row
        mom_row_num = max_row - 1
        yoy_row_num = max_row
        for row_idx in range(2, max_row + 1):
            period_cell = ws.cell(row=row_idx, column=1)
            period_cell.fill = period_fill
            period_cell.font = period_font
            period_cell.alignment = left_align
            period_cell.border = thin_border
            is_mom = row_idx == mom_row_num
            is_yoy = row_idx == yoy_row_num
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                if is_mom:
                    cell.fill = mom_fill
                    cell.number_format = '0.0%'
                elif is_yoy:
                    cell.fill = yoy_fill
                    cell.number_format = '0.0%'
                else:
                    col_header = ws.cell(row=1, column=col_idx).value
                    if 'CPC' in str(col_header) or 'CTR' in str(col_header):
                        cell.number_format = '#,##0.00'
                    elif 'Spend' in str(col_header):
                        cell.number_format = '$#,##0.00'
                    else:
                        cell.number_format = '#,##0'


def monthly_write_summaries_to_buffer(summary_tables):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for table_name, table_df in summary_tables.items():
            sheet_name = table_name[:31]
            table_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            num_data_rows = len(table_df)
            mom_row = num_data_rows
            yoy_row = num_data_rows + 1
            current_row = 2
            prev_row = 3
            year_ago_row = 4
            num_cols = len(table_df.columns)
            for col_idx in range(2, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{mom_row}'] = f'=IF(OR({col_letter}{prev_row}=0,{col_letter}{prev_row}=""),"-",({col_letter}{current_row}-{col_letter}{prev_row})/{col_letter}{prev_row})'
            for col_idx in range(2, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{yoy_row}'] = f'=IF(OR({col_letter}{year_ago_row}=0,{col_letter}{year_ago_row}=""),"-",({col_letter}{current_row}-{col_letter}{year_ago_row})/{col_letter}{year_ago_row})'
        monthly_format_summary_tables(writer.book, summary_tables.keys())
    buf.seek(0)
    return buf

# ---- Monthly insights ----

def monthly_generate_insights(df):
    if 'Month_dt' not in df.columns:
        df = monthly_prepare_dataframe(df)
    df_sorted = df.sort_values('Month_dt', ascending=False)
    unique_months = df_sorted.drop_duplicates(subset=['Month_dt'])[['Month', 'Month_dt']].reset_index(drop=True)
    if len(unique_months) < 2:
        return "ERROR: Need at least 2 months of data for MoM comparison"
    current_month_dt = unique_months.iloc[0]['Month_dt']
    prev_month_dt = unique_months.iloc[1]['Month_dt']
    current_month = unique_months.iloc[0]['Month']
    prev_month = unique_months.iloc[1]['Month']

    def calc_wow(current_df, prev_df, metric):
        curr_val = pd.to_numeric(current_df[metric], errors='coerce').fillna(0).sum() if metric in current_df.columns else 0
        prev_val = pd.to_numeric(prev_df[metric], errors='coerce').fillna(0).sum() if metric in prev_df.columns else 0
        if prev_val == 0:
            return 0, curr_val, prev_val
        return ((curr_val - prev_val) / prev_val) * 100, curr_val, prev_val

    def calc_rate(data_df, num_col, den_col, mult=1):
        num = pd.to_numeric(data_df[num_col], errors='coerce').fillna(0).sum() if num_col in data_df.columns else 0
        den = pd.to_numeric(data_df[den_col], errors='coerce').fillna(0).sum() if den_col in data_df.columns else 0
        return (num / den) * mult if den > 0 else 0

    def format_change(value):
        return f"{value:+.1f}%"

    def segment_insights(segment_name, curr_data, prev_data):
        if len(curr_data) == 0 or len(prev_data) == 0:
            return [f"{segment_name}: No data available"]
        clicks_change, _, _ = calc_wow(curr_data, prev_data, 'Clicks')
        spend_change, _, _ = calc_wow(curr_data, prev_data, 'Cost')
        cpc_curr = calc_rate(curr_data, 'Cost', 'Clicks', 1)
        cpc_prev = calc_rate(prev_data, 'Cost', 'Clicks', 1)
        cpc_change = ((cpc_curr - cpc_prev) / cpc_prev * 100) if cpc_prev > 0 else 0
        ctr_curr = calc_rate(curr_data, 'Clicks', 'Impr.', 100)
        ctr_prev = calc_rate(prev_data, 'Clicks', 'Impr.', 100)
        ctr_change = ((ctr_curr - ctr_prev) / ctr_prev * 100) if ctr_prev > 0 else 0
        ctr_direction = "increased" if ctr_change > 0 else "decreased"
        conversion_metrics = {
            'eCom Order - New': 'eCom Orders', 'Lead Form Submission - New': 'Lead Forms',
            'Address Capture': 'Address Captures', 'Begin Checkout': 'Begin Checkouts',
            'Quality Sales Call - AN': 'Quality Sales Calls', 'Main Sales Number': 'Main Sales Number',
            'Contact Us Page': 'Contact Us Page'
        }
        up_conv, down_conv = [], []
        for metric, display_name in conversion_metrics.items():
            if metric in curr_data.columns:
                change, _, _ = calc_wow(curr_data, prev_data, metric)
                if change > 0:
                    up_conv.append(display_name)
                elif change < 0:
                    down_conv.append(display_name)
        text = (f"{segment_name}: Clicks {format_change(clicks_change)} with spend {format_change(spend_change)} "
                f"with CPCs {format_change(cpc_change)}. "
                f"Average CTR {ctr_direction} {abs(ctr_change):.1f}% ({ctr_prev:.2f}% to {ctr_curr:.2f}%). ")
        if down_conv:
            text += f"All conversions increased MoM except for [{', '.join(down_conv)}]"
        elif up_conv:
            text += f"All conversion metrics flat or down MoM except for [{', '.join(up_conv)}]"
        else:
            text += "All conversion metrics relatively flat MoM"
        return [text]

    current_data = df_sorted[df_sorted['Month_dt'] == current_month_dt]
    prev_data = df_sorted[df_sorted['Month_dt'] == prev_month_dt]
    if 'Campaign Type' not in df.columns and 'Category (with Brand vs NB)' in df.columns:
        df['Campaign Type'] = df['Category (with Brand vs NB)'].apply(
            lambda x: 'Brand' if 'Brand' in str(x) and 'NB' not in str(x) else 'NonBrand')
        current_data = df_sorted[df_sorted['Month_dt'] == current_month_dt]
        prev_data = df_sorted[df_sorted['Month_dt'] == prev_month_dt]

    report = []
    report.append("=" * 80)
    report.append("MONTH-OVER-MONTH MARKETING PERFORMANCE REPORT")
    report.append(f"Comparing: {current_month} vs {prev_month}")
    report.append("=" * 80)
    report.append("")
    if 'Customer Type' not in df.columns:
        report.append("ERROR: 'Customer Type' column not found in data")
        return '\n'.join(report)
    for cust_type in ['CC', 'NC']:
        cust_curr = current_data[current_data['Customer Type'] == cust_type]
        cust_prev = prev_data[prev_data['Customer Type'] == cust_type]
        if len(cust_curr) == 0:
            continue
        report.append("-" * 80)
        report.append(f"{cust_type} - CURRENT CUSTOMERS" if cust_type == 'CC' else f"{cust_type} - NON CUSTOMERS")
        report.append("-" * 80)
        report.append("")
        report.extend(segment_insights(f"{cust_type} - Overall", cust_curr, cust_prev))
        report.append("")
        if 'Engine' in df.columns:
            for engine in sorted(cust_curr['Engine'].dropna().unique()):
                report.extend(segment_insights(f"{cust_type} - {engine}",
                    cust_curr[cust_curr['Engine'] == engine], cust_prev[cust_prev['Engine'] == engine]))
                report.append("")
        if 'Campaign Type' in df.columns:
            for camp_type in ['Brand', 'NonBrand']:
                camp_curr = cust_curr[cust_curr['Campaign Type'] == camp_type]
                camp_prev = cust_prev[cust_prev['Campaign Type'] == camp_type]
                if len(camp_curr) > 0:
                    report.extend(segment_insights(f"{cust_type} - {camp_type}", camp_curr, camp_prev))
                    report.append("")
        report.append("")
    report.append("=" * 80)
    report.append("KEY TAKEAWAYS")
    report.append("=" * 80)
    report.append("")
    oc, _, _ = calc_wow(current_data, prev_data, 'Clicks')
    os_, _, _ = calc_wow(current_data, prev_data, 'Cost')
    occ = calc_rate(current_data, 'Cost', 'Clicks', 1)
    ocp = calc_rate(prev_data, 'Cost', 'Clicks', 1)
    ocpc_change = ((occ - ocp) / ocp * 100) if ocp > 0 else 0
    report.append(f"* Overall Performance: Clicks {format_change(oc)}, Spend {format_change(os_)}, CPC {format_change(ocpc_change)}")
    if 'CC' in current_data['Customer Type'].values and 'NC' in current_data['Customer Type'].values:
        ccc, _, _ = calc_wow(current_data[current_data['Customer Type'] == 'CC'], prev_data[prev_data['Customer Type'] == 'CC'], 'Clicks')
        ncc, _, _ = calc_wow(current_data[current_data['Customer Type'] == 'NC'], prev_data[prev_data['Customer Type'] == 'NC'], 'Clicks')
        report.append(f"* CC vs NC: CC Clicks {format_change(ccc)}, NC Clicks {format_change(ncc)}")
    if 'Campaign Type' in df.columns:
        bc, _, _ = calc_wow(current_data[current_data['Campaign Type'] == 'Brand'], prev_data[prev_data['Campaign Type'] == 'Brand'], 'Clicks')
        nbc, _, _ = calc_wow(current_data[current_data['Campaign Type'] == 'NonBrand'], prev_data[prev_data['Campaign Type'] == 'NonBrand'], 'Clicks')
        report.append(f"* Brand vs NonBrand: Brand Clicks {format_change(bc)}, NonBrand Clicks {format_change(nbc)}")
    if 'Engine' in df.columns:
        for engine in ['Google', 'Bing']:
            if engine in current_data['Engine'].values:
                ec, _, _ = calc_wow(current_data[current_data['Engine'] == engine], prev_data[prev_data['Engine'] == engine], 'Clicks')
                report.append(f"* {engine}: Clicks {format_change(ec)}")
    report.append("")
    report.append("=" * 80)
    return '\n'.join(report)

# ---- Monthly UI ----

def run_monthly_report():
    st.header("Monthly Campaign Summary Report")
    st.markdown("Upload a monthly SA360 export (CSV or Excel) to generate summary tables and insights.")

    uploaded = st.file_uploader("Drop your file here", type=["csv", "xlsx", "xls"], key="monthly_upload")
    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    with st.spinner("Reading file..."):
        df, error = monthly_load_file(uploaded)
    if error:
        st.error(f"Could not parse file: {error}")
        return

    st.success(f"Loaded {len(df):,} rows from {uploaded.name}")
    df = monthly_prepare_dataframe(df)
    df = monthly_add_parsed_columns(df, campaign_col='Campaign')
    unique_months = df.sort_values('Month_dt', ascending=False).drop_duplicates(subset=['Month_dt'])

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Months in file", len(unique_months))
    with col2:
        st.metric("Total rows", f"{len(df):,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'Customer Type': 'Type', 'count': 'Rows'}), hide_index=True)
        with c2:
            st.markdown("**Engine**")
            st.dataframe(df['Engine'].value_counts().reset_index().rename(
                columns={'Engine': 'Type', 'count': 'Rows'}), hide_index=True)
        with c3:
            st.markdown("**Brand**")
            st.dataframe(df['Brand'].value_counts().reset_index().rename(
                columns={'Brand': 'Type', 'count': 'Rows'}), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    st.divider()
    if len(unique_months) < 2:
        st.warning(f"Need at least 2 months of data. Found {len(unique_months)} month(s).")
        return

    if st.button("Generate Report", type="primary", use_container_width=True, key="monthly_btn"):
        with st.spinner("Creating summary tables..."):
            summary_tables = monthly_create_formatted_summaries(df)
        if not summary_tables:
            st.error("No summary tables could be created. Check that Customer Type was parsed correctly.")
            return
        with st.spinner("Building Excel file..."):
            excel_buf = monthly_write_summaries_to_buffer(summary_tables)
        st.markdown("---")
        st.subheader("Downloads")
        st.download_button(
            label="Download Excel Summary (CB_MONTHLY_SUMMARY.xlsx)",
            data=excel_buf, file_name="CB_MONTHLY_SUMMARY.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
        with st.spinner("Generating insights..."):
            insights_text = monthly_generate_insights(df)
        st.download_button(
            label="Download Insights (Monthly_summary_insights.txt)",
            data=insights_text.encode('utf-8'), file_name="Monthly_summary_insights.txt",
            mime="text/plain", use_container_width=True)
        with st.expander("Preview insights", expanded=False):
            st.text(insights_text)
        st.success(f"Done. {len(summary_tables)} summary table(s) generated.")


# =============================================================================
#
#   KEYWORD ANALYSIS
#
# =============================================================================

# ---- Keyword parsing (reuses monthly parsing) ----

def keyword_add_parsed_columns(df, campaign_col='Campaign'):
    df = df.copy()
    parsed = df[campaign_col].apply(monthly_parse_campaign_name)
    df['Customer Type'] = parsed.apply(lambda x: x['Customer_type'])
    df['Engine'] = parsed.apply(lambda x: x['engine'])
    df['Brand'] = parsed.apply(lambda x: x['brand'])
    unparsed_ctype = df[df['Customer Type'].isna()]
    if len(unparsed_ctype) > 0:
        st.warning(f"Could not parse Customer Type from {len(unparsed_ctype)} campaigns")
    unparsed_engine = df[df['Engine'].isna()]
    if len(unparsed_engine) > 0:
        st.warning(f"Could not parse Engine from {len(unparsed_engine)} campaigns")
    unparsed_brand = df[df['Brand'].isna()]
    if len(unparsed_brand) > 0:
        st.warning(f"Could not parse Brand from {len(unparsed_brand)} campaigns")
    return df

# ---- Keyword file loading ----

def keyword_load_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        for enc in ("utf-8", "utf-8-sig", "utf-16"):
            for skip in [0, 1, 2, 3]:
                try:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, encoding=enc, sep=None, engine='python', skiprows=skip if skip > 0 else None)
                    if 'Campaign' in df.columns and 'Search keyword' in df.columns:
                        return df, None
                except Exception:
                    continue
        return None, "Could not parse CSV. Ensure it contains 'Campaign' and 'Search keyword' columns."
    elif name.endswith(('.xlsx', '.xls')):
        for skip in [0, 1, 2, 3]:
            try:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, skiprows=skip)
                if 'Campaign' in df.columns and 'Search keyword' in df.columns:
                    return df, None
            except Exception:
                continue
        return None, "Could not parse Excel file. Ensure it contains 'Campaign' and 'Search keyword' columns."
    return None, f"Unsupported file type: {name}"

# ---- Keyword analysis ----

def keyword_analyze(df, sort_by='Clicks'):
    if 'Campaign' in df.columns:
        df = keyword_add_parsed_columns(df, campaign_col='Campaign')
    required = ['Search keyword', 'Customer Type', 'Brand']
    missing = [col for col in required if col not in df.columns]
    if missing:
        return None, [f"Missing required columns: {missing}"]
    metrics = ['Clicks', 'Cost', 'Visits']
    agg_metrics = [m for m in metrics if m in df.columns]
    for col in agg_metrics:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
    grouped = df.groupby(['Customer Type', 'Brand', 'Search keyword'], as_index=False).agg(
        {m: 'sum' for m in agg_metrics})
    combos = grouped[['Customer Type', 'Brand']].drop_duplicates().sort_values(
        ['Customer Type', 'Brand']).values
    results = {}
    messages = []
    for cust_type, brand_type in combos:
        subset = grouped[(grouped['Customer Type'] == cust_type) & (grouped['Brand'] == brand_type)].copy()
        sort_col = sort_by if sort_by in subset.columns else 'Clicks'
        top_10 = subset.nlargest(10, sort_col)
        display_cols = [c for c in ['Search keyword', 'Clicks', 'Cost'] if c in top_10.columns]
        top_10_display = top_10[display_cols].reset_index(drop=True)
        top_10_display.index = top_10_display.index + 1
        key = f"{cust_type}_{brand_type}".replace(' ', '')
        results[key] = {
            'data': top_10_display, 'customer_type': cust_type,
            'brand_type': brand_type, 'sort_metric': sort_col, 'total_keywords': len(subset)
        }
        messages.append(f"{cust_type} - {brand_type}: {len(subset)} total keywords, top 10 by {sort_col}")
    return results, messages

# ---- Keyword Excel output ----

def keyword_create_excel(results):
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for key, result in sorted(results.items()):
        ws = wb.create_sheet(key[:31])
        ws['A1'] = f"Top 10 Keywords: {result['customer_type']} - {result['brand_type']}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A2'] = f"Total Keywords in Segment: {result['total_keywords']}"
        ws['A2'].font = Font(italic=True, size=10)
        df = result['data'].reset_index()
        df.columns = ['Rank'] + list(result['data'].columns)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=False)
                if col_idx > 2 and isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 12
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---- Keyword UI ----

def run_keyword_analysis():
    st.header("Keyword Analysis Tool")
    st.markdown("Upload an SA360 keyword export (CSV or Excel) to generate top 10 keyword tables by Customer Type and Brand/NB.")

    uploaded = st.file_uploader("Drop your file here", type=["csv", "xlsx", "xls"], key="keyword_upload")
    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    with st.spinner("Reading file..."):
        df, error = keyword_load_file(uploaded)
    if error:
        st.error(f"Could not parse file: {error}")
        return

    st.success(f"Loaded {len(df):,} rows from {uploaded.name}")

    if 'Campaign' in df.columns:
        df = keyword_add_parsed_columns(df, campaign_col='Campaign')

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total rows", f"{len(df):,}")
    with col2:
        st.metric("Unique keywords", f"{df['Search keyword'].nunique():,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'Customer Type': 'Type', 'count': 'Rows'}), hide_index=True)
        with c2:
            st.markdown("**Engine**")
            st.dataframe(df['Engine'].value_counts().reset_index().rename(
                columns={'Engine': 'Type', 'count': 'Rows'}), hide_index=True)
        with c3:
            st.markdown("**Brand**")
            st.dataframe(df['Brand'].value_counts().reset_index().rename(
                columns={'Brand': 'Type', 'count': 'Rows'}), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    st.divider()
    available_metrics = [m for m in ['Clicks', 'Cost', 'Visits'] if m in df.columns]
    sort_by = st.selectbox("Sort top keywords by:", available_metrics, index=0)

    if st.button("Generate Analysis", type="primary", use_container_width=True, key="keyword_btn"):
        with st.spinner("Analyzing keywords..."):
            results, messages = keyword_analyze(df, sort_by=sort_by)
        if results is None:
            for msg in messages:
                st.error(msg)
            return
        for key, result in sorted(results.items()):
            with st.expander(
                f"{result['customer_type']} - {result['brand_type']}  "
                f"({result['total_keywords']} keywords, top 10 by {result['sort_metric']})",
                expanded=True
            ):
                st.dataframe(result['data'], use_container_width=True, hide_index=False)
        st.markdown("---")
        with st.spinner("Building Excel file..."):
            excel_buf = keyword_create_excel(results)
        st.download_button(
            label="Download keyword_analysis_results.xlsx",
            data=excel_buf, file_name="keyword_analysis_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
        st.success(f"Done. {len(results)} segment(s) analyzed.")


# =============================================================================
#
#   MAIN APP
#
# =============================================================================

def main():
    st.set_page_config(page_title="CB Reporting Hub", layout="wide")

    st.title("CB Reporting Hub")

    report_type = st.selectbox(
        "Select a report to run:",
        ["Weekly WoW Report", "Monthly Campaign Summary", "Keyword Analysis"]
    )

    st.divider()

    if report_type == "Weekly WoW Report":
        run_weekly_report()
    elif report_type == "Monthly Campaign Summary":
        run_monthly_report()
    elif report_type == "Keyword Analysis":
        run_keyword_analysis()


if __name__ == "__main__":
    main()
