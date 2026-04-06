#!/usr/bin/env python3
"""
Monthly Campaign Summary Tables with Automated Insights — Streamlit App
Processes SA360 monthly data and generates formatted Excel summaries + text insights.

Usage: streamlit run monthly_report_app.py
"""

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from datetime import datetime

# =============================================================================
# CAMPAIGN NAME PARSING
# =============================================================================

def parse_campaign_name(campaign_name):
    """
    Parse campaign name to extract customer type, engine, and brand.
    Parts: [0]Brand [1]Channel [2]Type [3]CustomerType [4-6]...details... [last]Engine
    
    Returns: dict with 'customer_type', 'engine', 'brand'
    """
    parts = campaign_name.split('_')
    
    result = {
        'Customer_type': None,
        'engine': None,
        'brand': None
    }
    
    # Engine: Search through all parts for Google or Bing
    for part in parts:
        part_clean = part.strip().lower()
        if 'google' in part_clean:
            result['engine'] = 'Google'
            break
        elif 'bing' in part_clean:
            result['engine'] = 'Bing'
            break
    
    # Customer Type: Look for CC or NC (typically 4th position)
    for part in parts:
        part_upper = part.upper()
        if part_upper in ['CC', 'NC']:
            result['Customer_type'] = part_upper
            break
    
    # Brand: Look for Brand, Br, Nonbr, or NonBrand patterns
    for i, part in enumerate(parts):
        part_lower = part.lower()
        if part_lower == 'nonbr' or part_lower == 'nonbrand':
            result['brand'] = 'NonBrand'
            break
        elif part_lower == 'br' or part_lower == 'brand':
            result['brand'] = 'Brand'
            break
    
    return result


def add_parsed_columns(df, campaign_col='Campaign'):
    """
    Add Customer Type, Engine, and Brand columns by parsing the campaign name.
    """
    df = df.copy()
    
    parsed = df[campaign_col].apply(parse_campaign_name)
    
    df['Customer Type'] = parsed.apply(lambda x: x['Customer_type'])
    df['Engine'] = parsed.apply(lambda x: x['engine'])
    df['Brand'] = parsed.apply(lambda x: x['brand'])
    
    unparsed_ctype = df[df['Customer Type'].isna()]
    if len(unparsed_ctype) > 0:
        st.warning(f"Could not parse Customer Type from {len(unparsed_ctype)} campaigns")
    
    unparsed_engine = df[df['Engine'].isna()]
    if len(unparsed_engine) > 0:
        st.warning(f"Could not parse Engine from {len(unparsed_engine)} campaigns")
    
    unparsed_brand = df[df['Brand'].isna()]
    if len(unparsed_brand) > 0:
        st.warning(f"Could not parse Brand from {len(unparsed_brand)} campaigns")
    
    return df

# =============================================================================
# DATA PREPARATION
# =============================================================================

def prepare_dataframe(df):
    """
    Prepare dataframe with proper datetime conversion for calendar-correct sorting
    """
    if pd.api.types.is_datetime64_any_dtype(df['Month']):
        df['Month_dt'] = pd.to_datetime(df['Month'])
    else:
        formats = [
            ('%y-%b', 'YY-Mon (e.g., 25-Aug)'),
            ('%b-%y', 'Mon-YY (e.g., Aug-25)'),
            ('%B %Y', 'Month YYYY (e.g., August 2025)'),
        ]
        
        parsed = False
        for fmt, desc in formats:
            attempt = pd.to_datetime(df['Month'], format=fmt, errors='coerce')
            success = attempt.notna().sum()
            if success > 0:
                df['Month_dt'] = attempt
                parsed = True
                break
        
        if not parsed:
            df['Month_dt'] = pd.to_datetime(df['Month'], errors='coerce')
    
    return df

# =============================================================================
# FILE LOADING
# =============================================================================

def load_file(uploaded_file):
    """Load CSV or Excel file and return dataframe."""
    name = uploaded_file.name.lower()

    if name.endswith('.csv'):
        for enc in ("utf-8-sig", "utf-8", "utf-16"):
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, encoding=enc, skiprows=2)
                if 'Campaign' in df.columns and 'Month' in df.columns:
                    return df, None
            except (UnicodeDecodeError, Exception):
                continue

        # Try without skiprows
        for enc in ("utf-8-sig", "utf-8", "utf-16"):
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, encoding=enc)
                if 'Campaign' in df.columns and 'Month' in df.columns:
                    return df, None
            except (UnicodeDecodeError, Exception):
                continue

        return None, "Could not parse CSV. Ensure it contains 'Campaign' and 'Month' columns."

    elif name.endswith(('.xlsx', '.xls')):
        for skip in [0, 2, 1, 3]:
            try:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, skiprows=skip)
                if 'Campaign' in df.columns and 'Month' in df.columns:
                    return df, None
            except Exception:
                continue

        return None, "Could not parse Excel file. Ensure it contains 'Campaign' and 'Month' columns."

    return None, f"Unsupported file type: {name}"

# =============================================================================
# SUMMARY TABLE CREATION
# =============================================================================

def create_formatted_summaries(df, available_metrics):
    """
    Create formatted summary tables for reporting
    Structure: CC/NC -> Overall, Google, Bing, Brand, NonBrand
    Uses Month_dt for correct calendar sorting
    """
    summary_tables = {}
    
    if 'Month_dt' not in df.columns:
        df = prepare_dataframe(df)
    
    df_sorted = df.sort_values('Month_dt', ascending=False)
    unique_months = df_sorted.drop_duplicates(subset=['Month_dt'])[['Month', 'Month_dt']].reset_index(drop=True)
    
    if len(unique_months) < 2:
        return summary_tables
    
    current_month_dt = unique_months.iloc[0]['Month_dt']
    current_month = unique_months.iloc[0]['Month']
    
    prev_month_dt = current_month_dt - pd.DateOffset(months=1)
    prev_month_match = df_sorted[df_sorted['Month_dt'] <= prev_month_dt].drop_duplicates(subset=['Month_dt'])
    if len(prev_month_match) > 0:
        prev_month_dt = prev_month_match.iloc[0]['Month_dt']
        prev_month = prev_month_match.iloc[0]['Month']
    else:
        if len(unique_months) >= 2:
            prev_month_dt = unique_months.iloc[1]['Month_dt']
            prev_month = unique_months.iloc[1]['Month']
        else:
            prev_month = None
            prev_month_dt = None
    
    year_ago_dt = current_month_dt - pd.DateOffset(years=1)
    year_ago_matches = df_sorted[
        (df_sorted['Month_dt'] >= year_ago_dt - pd.Timedelta(days=45)) &
        (df_sorted['Month_dt'] <= year_ago_dt + pd.Timedelta(days=45))
    ].drop_duplicates(subset=['Month_dt'])
    
    if len(year_ago_matches) > 0:
        year_ago_matches = year_ago_matches.copy()
        year_ago_matches['diff'] = abs((year_ago_matches['Month_dt'] - year_ago_dt).dt.days)
        closest = year_ago_matches.sort_values('diff').iloc[0]
        year_ago_month_dt = closest['Month_dt']
        year_ago_month = closest['Month']
    else:
        year_ago_month = None
        year_ago_month_dt = None
    
    has_customer_type = 'Customer Type' in df.columns
    has_engine = 'Engine' in df.columns
    has_campaign_type = 'Brand' in df.columns or 'Category (with Brand vs NB)' in df.columns
    
    if 'Category (with Brand vs NB)' in df.columns and 'Campaign Type' not in df.columns:
        df['Campaign Type'] = df['Category (with Brand vs NB)'].apply(
            lambda x: 'Brand' if 'Brand' in str(x) and 'NB' not in str(x) else 'NonBrand'
        )
        has_campaign_type = True
    
    if not has_customer_type:
        return summary_tables
    
    customer_types = df['Customer Type'].dropna().unique()
    
    for cust_type in customer_types:
        summary_tables[f'{cust_type} - Overall'] = create_summary_table(
            df, cust_type, None, None,
            current_month_dt, prev_month_dt, year_ago_month_dt
        )
        
        if has_engine:
            engines = df[df['Customer Type'] == cust_type]['Engine'].dropna().unique()
            for engine in engines:
                summary_tables[f'{cust_type} - {engine}'] = create_summary_table(
                    df, cust_type, 'Engine', engine,
                    current_month_dt, prev_month_dt, year_ago_month_dt
                )
        
        if has_campaign_type:
            campaign_types = df[df['Customer Type'] == cust_type]['Brand'].dropna().unique()
            for camp_type in campaign_types:
                summary_tables[f'{cust_type} - {camp_type}'] = create_summary_table(
                    df, cust_type, 'Brand', camp_type,
                    current_month_dt, prev_month_dt, year_ago_month_dt
                )
    
    return summary_tables


def create_summary_table(df, customer_type, filter_col, filter_value,
                         current_month_dt, prev_month_dt, year_ago_month_dt):
    """
    Create a summary table using datetime for filtering
    Returns display values (Nov-25) not datetime
    """
    summary_metrics = [
        ('Spend', 'Cost'),
        ('Impressions', 'Impr.'),
        ('Clicks', 'Clicks'),
        ('avg CPC', None),
        ('avg CTR', None),
        ('eCom Orders', 'CB eCom Order Tag - New'),
        ('Lead Form', 'CB General Lead Form Submission - New'),
        ('Address Capture', 'Address Capture'),
        ('Begin Checkout', 'Begin Checkout'),
        ('Total Conversions - VBB', 'Total Conversions - VBB'),
        ('Chat Initiation - Order Services', 'Chat Initiation - Order Services'),
        ('Quality Sales Calls (Offline)', 'Quality Sales Call - AN'),
        ('Main Sales Number', 'Main Sales Number'),
        ('Contact Us Page', 'Contact Us Page')
    ]
    
    df_filtered = df[df['Customer Type'] == customer_type].copy()
    
    if filter_col and filter_value:
        df_filtered = df_filtered[df_filtered[filter_col] == filter_value]
    
    if filter_col is None:
        table_name = f"{customer_type} - Overall"
    else:
        table_name = f"{customer_type} - {filter_value}"
    
    data = {table_name: []}
    for display_name, _ in summary_metrics:
        data[display_name] = []
    
    periods = [
        ('current', current_month_dt),
        ('previous', prev_month_dt),
        ('year_ago', year_ago_month_dt)
    ]
    
    for period_name, month_dt in periods:
        if month_dt is None:
            continue
        
        month_data = df_filtered[df_filtered['Month_dt'] == month_dt]
        
        if len(month_data) == 0:
            continue
        
        month_display = month_data['Month'].iloc[0]
        
        if isinstance(month_display, pd.Timestamp):
            month_display = month_display.strftime('%b-%y')
        
        data[table_name].append(str(month_display))
        
        for display_name, actual_col in summary_metrics:
            if display_name == 'avg CPC':
                total_cost = data['Spend'][-1] if 'Spend' in data and len(data['Spend']) > 0 else 0
                total_clicks = data['Clicks'][-1] if 'Clicks' in data and len(data['Clicks']) > 0 else 0
                avg_cpc = total_cost / total_clicks if total_clicks > 0 else 0
                data[display_name].append(round(avg_cpc, 2))
            elif display_name == 'avg CTR':
                total_impr = data['Impressions'][-1] if 'Impressions' in data and len(data['Impressions']) > 0 else 0
                total_clicks = data['Clicks'][-1] if 'Clicks' in data and len(data['Clicks']) > 0 else 0
                avg_ctr = (total_clicks / total_impr * 100) if total_impr > 0 else 0
                data[display_name].append(round(avg_ctr, 2))
            elif actual_col in month_data.columns:
                value = month_data[actual_col].sum()
                try:
                    value = pd.to_numeric(value, errors='coerce')
                    data[display_name].append(round(value, 2) if pd.notna(value) else 0)
                except (ValueError, TypeError):
                    data[display_name].append(0)
            else:
                data[display_name].append(0)
    
    data[table_name].append('MoM')
    data[table_name].append('YoY')
    
    for display_name, _ in summary_metrics:
        data[display_name].append(None)
        data[display_name].append(None)
    
    summary_df = pd.DataFrame(data)
    
    return summary_df

# =============================================================================
# EXCEL OUTPUT AND FORMATTING
# =============================================================================

def write_summaries_to_buffer(summary_tables):
    """
    Write summary tables to an in-memory Excel buffer with formula-based MoM and YoY.
    Returns a BytesIO buffer.
    """
    buf = BytesIO()

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for table_name, table_df in summary_tables.items():
            sheet_name = table_name[:31]
            table_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            num_data_rows = len(table_df)
            mom_row = num_data_rows
            yoy_row = num_data_rows + 1

            current_row = 2
            prev_row = 3
            year_ago_row = 4

            num_cols = len(table_df.columns)

            for col_idx in range(2, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                mom_formula = f'=IF(OR({col_letter}{prev_row}=0,{col_letter}{prev_row}=""),"-",({col_letter}{current_row}-{col_letter}{prev_row})/{col_letter}{prev_row})'
                ws[f'{col_letter}{mom_row}'] = mom_formula

            for col_idx in range(2, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                yoy_formula = f'=IF(OR({col_letter}{year_ago_row}=0,{col_letter}{year_ago_row}=""),"-",({col_letter}{current_row}-{col_letter}{year_ago_row})/{col_letter}{year_ago_row})'
                ws[f'{col_letter}{yoy_row}'] = yoy_formula

        workbook = writer.book
        format_summary_tables(workbook, summary_tables.keys())

    buf.seek(0)
    return buf


def format_summary_tables(workbook, sheet_names):
    """
    Apply formatting to summary tables
    """
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    
    period_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    period_font = Font(bold=True)
    
    mom_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    yoy_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in sheet_names:
        sheet_name_actual = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
        
        if sheet_name_actual not in workbook.sheetnames:
            continue
        
        ws = workbook[sheet_name_actual]
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        max_row = ws.max_row
        mom_row_num = max_row - 1
        yoy_row_num = max_row
        
        for row_idx in range(2, max_row + 1):
            period_cell = ws.cell(row=row_idx, column=1)
            
            period_cell.fill = period_fill
            period_cell.font = period_font
            period_cell.alignment = left_align
            period_cell.border = thin_border
            
            is_mom = row_idx == mom_row_num
            is_yoy = row_idx == yoy_row_num
            
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                
                if is_mom:
                    cell.fill = mom_fill
                    cell.number_format = '0.0%'
                elif is_yoy:
                    cell.fill = yoy_fill
                    cell.number_format = '0.0%'
                else:
                    col_header = ws.cell(row=1, column=col_idx).value
                    if 'CPC' in str(col_header) or 'CTR' in str(col_header):
                        cell.number_format = '#,##0.00'
                    elif 'Spend' in str(col_header):
                        cell.number_format = '$#,##0.00'
                    else:
                        cell.number_format = '#,##0'

# =============================================================================
# INSIGHTS GENERATION
# =============================================================================

def generate_summary_insights(df):
    """
    Generate automated summary insights matching the summary tables structure.
    Returns the report as a string.
    """
    if 'Month_dt' not in df.columns:
        df = prepare_dataframe(df)
    
    df_sorted = df.sort_values('Month_dt', ascending=False)
    unique_months = df_sorted.drop_duplicates(subset=['Month_dt'])[['Month', 'Month_dt']].reset_index(drop=True)
    
    if len(unique_months) < 2:
        return "ERROR: Need at least 2 months of data for MoM comparison"
    
    current_month_dt = unique_months.iloc[0]['Month_dt']
    prev_month_dt = unique_months.iloc[1]['Month_dt']
    
    current_month = unique_months.iloc[0]['Month']
    prev_month = unique_months.iloc[1]['Month']
    
    def calc_wow(current_df, prev_df, metric):
        if metric in current_df.columns:
            curr_val = pd.to_numeric(current_df[metric], errors='coerce').fillna(0).sum()
        else:
            curr_val = 0
        if metric in prev_df.columns:
            prev_val = pd.to_numeric(prev_df[metric], errors='coerce').fillna(0).sum()
        else:
            prev_val = 0
        if prev_val == 0:
            return 0, curr_val, prev_val
        pct_change = ((curr_val - prev_val) / prev_val) * 100
        return pct_change, curr_val, prev_val
    
    def calc_rate(data_df, numerator_col, denominator_col, multiplier=1):
        if numerator_col in data_df.columns:
            numerator = pd.to_numeric(data_df[numerator_col], errors='coerce').fillna(0).sum()
        else:
            numerator = 0
        if denominator_col in data_df.columns:
            denominator = pd.to_numeric(data_df[denominator_col], errors='coerce').fillna(0).sum()
        else:
            denominator = 0
        if denominator == 0:
            return 0
        return (numerator / denominator) * multiplier
    
    def format_change(value):
        return f"{value:+.1f}%"
    
    def generate_segment_insights(segment_name, curr_data, prev_data):
        insights = []
        
        if len(curr_data) == 0 or len(prev_data) == 0:
            insights.append(f"{segment_name}: No data available")
            return insights
        
        clicks_change, clicks_curr, clicks_prev = calc_wow(curr_data, prev_data, 'Clicks')
        spend_change, spend_curr, spend_prev = calc_wow(curr_data, prev_data, 'Cost')
        
        cpc_curr = calc_rate(curr_data, 'Cost', 'Clicks', 1)
        cpc_prev = calc_rate(prev_data, 'Cost', 'Clicks', 1)
        cpc_change = ((cpc_curr - cpc_prev) / cpc_prev * 100) if cpc_prev > 0 else 0
        
        ctr_curr = calc_rate(curr_data, 'Clicks', 'Impr.', 100)
        ctr_prev = calc_rate(prev_data, 'Clicks', 'Impr.', 100)
        ctr_change = ((ctr_curr - ctr_prev) / ctr_prev * 100) if ctr_prev > 0 else 0
        ctr_direction = "increased" if ctr_change > 0 else "decreased"
        
        conversion_metrics = {
            'eCom Order - New': 'eCom Orders',
            'Lead Form Submission - New': 'Lead Forms',
            'Address Capture': 'Address Captures',
            'Begin Checkout': 'Begin Checkouts',
            'Quality Sales Call - AN': 'Quality Sales Calls',
            'Main Sales Number': 'Main Sales Number',
            'Contact Us Page': 'Contact Us Page'
        }
        
        up_conversions = []
        down_conversions = []
        
        for metric, display_name in conversion_metrics.items():
            if metric in curr_data.columns:
                change, _, _ = calc_wow(curr_data, prev_data, metric)
                if change > 0:
                    up_conversions.append(display_name)
                elif change < 0:
                    down_conversions.append(display_name)
        
        insight_text = (f"{segment_name}: Clicks {format_change(clicks_change)} with spend {format_change(spend_change)} "
                       f"with CPCs {format_change(cpc_change)}. ")
        
        insight_text += f"Average CTR {ctr_direction} {abs(ctr_change):.1f}% ({ctr_prev:.2f}% to {ctr_curr:.2f}%). "
        
        if down_conversions:
            except_text = ', '.join(down_conversions)
            insight_text += f"All conversions increased MoM except for [{except_text}]"
        elif up_conversions:
            except_text = ', '.join(up_conversions)
            insight_text += f"All conversion metrics flat or down MoM except for [{except_text}]"
        else:
            insight_text += "All conversion metrics relatively flat MoM"
        
        insights.append(insight_text)
        return insights
    
    current_data = df_sorted[df_sorted['Month_dt'] == current_month_dt]
    prev_data = df_sorted[df_sorted['Month_dt'] == prev_month_dt]
    
    if 'Campaign Type' not in df.columns and 'Category (with Brand vs NB)' in df.columns:
        df['Campaign Type'] = df['Category (with Brand vs NB)'].apply(
            lambda x: 'Brand' if 'Brand' in str(x) and 'NB' not in str(x) else 'NonBrand'
        )
        current_data = df_sorted[df_sorted['Month_dt'] == current_month_dt]
        prev_data = df_sorted[df_sorted['Month_dt'] == prev_month_dt]
    
    report = []
    report.append("=" * 80)
    report.append(f"MONTH-OVER-MONTH MARKETING PERFORMANCE REPORT")
    report.append(f"Comparing: {current_month} vs {prev_month}")
    report.append("=" * 80)
    report.append("")
    
    if 'Customer Type' not in df.columns:
        report.append("ERROR: 'Customer Type' column not found in data")
        return '\n'.join(report)
    
    customer_types = ['CC', 'NC']
    
    for cust_type in customer_types:
        cust_curr = current_data[current_data['Customer Type'] == cust_type]
        cust_prev = prev_data[prev_data['Customer Type'] == cust_type]
        
        if len(cust_curr) == 0:
            continue
        
        report.append("-" * 80)
        report.append(f"{cust_type} - CURRENT CUSTOMERS" if cust_type == 'CC' else f"{cust_type} - NON CUSTOMERS")
        report.append("-" * 80)
        report.append("")
        
        insights = generate_segment_insights(f"{cust_type} - Overall", cust_curr, cust_prev)
        report.extend(insights)
        report.append("")
        
        if 'Engine' in df.columns:
            engines = cust_curr['Engine'].dropna().unique()
            for engine in sorted(engines):
                engine_curr = cust_curr[cust_curr['Engine'] == engine]
                engine_prev = cust_prev[cust_prev['Engine'] == engine]
                insights = generate_segment_insights(f"{cust_type} - {engine}", engine_curr, engine_prev)
                report.extend(insights)
                report.append("")
        
        if 'Campaign Type' in df.columns:
            for camp_type in ['Brand', 'NonBrand']:
                camp_curr = cust_curr[cust_curr['Campaign Type'] == camp_type]
                camp_prev = cust_prev[cust_prev['Campaign Type'] == camp_type]
                if len(camp_curr) > 0:
                    insights = generate_segment_insights(f"{cust_type} - {camp_type}", camp_curr, camp_prev)
                    report.extend(insights)
                    report.append("")
        
        report.append("")
    
    report.append("=" * 80)
    report.append("KEY TAKEAWAYS")
    report.append("=" * 80)
    report.append("")
    
    overall_clicks_change, _, _ = calc_wow(current_data, prev_data, 'Clicks')
    overall_spend_change, _, _ = calc_wow(current_data, prev_data, 'Cost')
    
    overall_cpc_curr = calc_rate(current_data, 'Cost', 'Clicks', 1)
    overall_cpc_prev = calc_rate(prev_data, 'Cost', 'Clicks', 1)
    overall_cpc_change = ((overall_cpc_curr - overall_cpc_prev) / overall_cpc_prev * 100) if overall_cpc_prev > 0 else 0
    
    report.append(f"* Overall Performance: Clicks {format_change(overall_clicks_change)}, "
                  f"Spend {format_change(overall_spend_change)}, "
                  f"CPC {format_change(overall_cpc_change)}")
    
    if 'CC' in current_data['Customer Type'].values and 'NC' in current_data['Customer Type'].values:
        cc_clicks, _, _ = calc_wow(
            current_data[current_data['Customer Type'] == 'CC'],
            prev_data[prev_data['Customer Type'] == 'CC'],
            'Clicks'
        )
        nc_clicks, _, _ = calc_wow(
            current_data[current_data['Customer Type'] == 'NC'],
            prev_data[prev_data['Customer Type'] == 'NC'],
            'Clicks'
        )
        report.append(f"* CC vs NC: CC Clicks {format_change(cc_clicks)}, NC Clicks {format_change(nc_clicks)}")
    
    if 'Campaign Type' in df.columns:
        brand_clicks, _, _ = calc_wow(
            current_data[current_data['Campaign Type'] == 'Brand'],
            prev_data[prev_data['Campaign Type'] == 'Brand'],
            'Clicks'
        )
        nb_clicks, _, _ = calc_wow(
            current_data[current_data['Campaign Type'] == 'NonBrand'],
            prev_data[prev_data['Campaign Type'] == 'NonBrand'],
            'Clicks'
        )
        report.append(f"* Brand vs NonBrand: Brand Clicks {format_change(brand_clicks)}, "
                      f"NonBrand Clicks {format_change(nb_clicks)}")
    
    if 'Engine' in df.columns:
        for engine in ['Google', 'Bing']:
            if engine in current_data['Engine'].values:
                engine_clicks, _, _ = calc_wow(
                    current_data[current_data['Engine'] == engine],
                    prev_data[prev_data['Engine'] == engine],
                    'Clicks'
                )
                report.append(f"* {engine}: Clicks {format_change(engine_clicks)}")
    
    report.append("")
    report.append("=" * 80)
    
    return '\n'.join(report)

# =============================================================================
# STREAMLIT APP
# =============================================================================

def main():
    st.set_page_config(page_title="Monthly Campaign Summary", layout="wide")

    st.title("Monthly Campaign Summary Report")
    st.markdown("Upload a monthly SA360 export (CSV or Excel) to generate summary tables and insights.")

    uploaded = st.file_uploader(
        "Drop your file here",
        type=["csv", "xlsx", "xls"],
        help="Accepts SA360 CSV or Excel exports with Campaign and Month columns"
    )

    if uploaded is None:
        st.info("Upload a file to get started.")
        return

    # ---- Load ----
    with st.spinner("Reading file..."):
        df, error = load_file(uploaded)

    if error:
        st.error(f"Could not parse file: {error}")
        return

    st.success(f"Loaded {len(df):,} rows from {uploaded.name}")

    # ---- Prepare ----
    df = prepare_dataframe(df)
    df = add_parsed_columns(df, campaign_col='Campaign')

    # ---- Summary info ----
    unique_months = df.sort_values('Month_dt', ascending=False).drop_duplicates(subset=['Month_dt'])

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Months in file", len(unique_months))
    with col2:
        st.metric("Total rows", f"{len(df):,}")
    with col3:
        st.metric("Campaigns", f"{df['Campaign'].nunique():,}")

    with st.expander("Classifications breakdown", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Customer Type**")
            st.dataframe(df['Customer Type'].value_counts().reset_index().rename(
                columns={'Customer Type': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c2:
            st.markdown("**Engine**")
            st.dataframe(df['Engine'].value_counts().reset_index().rename(
                columns={'Engine': 'Type', 'count': 'Rows'}
            ), hide_index=True)
        with c3:
            st.markdown("**Brand**")
            st.dataframe(df['Brand'].value_counts().reset_index().rename(
                columns={'Brand': 'Type', 'count': 'Rows'}
            ), hide_index=True)

    with st.expander("Preview raw data", expanded=False):
        st.dataframe(df.head(50), use_container_width=True)

    # ---- Generate ----
    st.divider()

    if len(unique_months) < 2:
        st.warning(f"Need at least 2 months of data. Found {len(unique_months)} month(s).")
        return

    if st.button("Generate Report", type="primary", use_container_width=True):

        # Excel summary
        with st.spinner("Creating summary tables..."):
            summary_tables = create_formatted_summaries(df, available_metrics=[])

        if not summary_tables:
            st.error("No summary tables could be created. Check that Customer Type was parsed correctly.")
            return

        with st.spinner("Building Excel file..."):
            excel_buf = write_summaries_to_buffer(summary_tables)

        st.markdown("---")
        st.subheader("Downloads")

        st.download_button(
            label="Download Excel Summary (CB_MONTHLY_SUMMARY.xlsx)",
            data=excel_buf,
            file_name="CB_MONTHLY_SUMMARY.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        # Text insights
        with st.spinner("Generating insights..."):
            insights_text = generate_summary_insights(df)

        st.download_button(
            label="Download Insights (Monthly_summary_insights.txt)",
            data=insights_text.encode('utf-8'),
            file_name="Monthly_summary_insights.txt",
            mime="text/plain",
            use_container_width=True,
        )

        with st.expander("Preview insights", expanded=False):
            st.text(insights_text)

        st.success(f"Done. {len(summary_tables)} summary table(s) generated.")


if __name__ == "__main__":
    main()
